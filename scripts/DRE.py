import pyodbc
import json
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfutils
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from collections import OrderedDict
# Importação necessária para o Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import file_renamer

# --- 1. CONFIGURAÇÃO E CONSULTAS SQL ---

CONN_STR = (
    "DRIVER={SQL Anywhere 17};"
    "HOST=NOTE-GO-273.go.local:2638;"
    "DBN=contabil;"
    "UID=ESTATISTICA002;"
    "PWD=U0T/wq6OdZ0oYSpvJRWGfg==;"
)

# Dicionário de traduções para textos fixos da interface
TRANSLATIONS = {
    'pt': {
        'report_title': 'DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO',
        'company': 'Empresa:',
        'cnpj': 'C.N.P.J.:',
        'period': 'Período:',
        'page': 'Folha:',
        'book_number': 'Número livro:',
        'emission': 'Emissão:',
        'time': 'Hora:',
        'description': 'Descrição',
        'balance': 'Saldo',
        'total': 'Total',
        'administrator': 'ADMINISTRADOR',
        'final_result': 'RESULTADO DO EXERCÍCIO'
    },
    'en': {
        'report_title': 'P&L',
        'company': 'Company:',
        'cnpj': 'C.N.P.J.:',  # ← Mantém C.N.P.J mesmo em inglês
        'period': 'Period:',
        'page': 'Page:',
        'book_number': 'Book number:',
        'emission': 'Date:',  # ← Mudança: Issue date → Date
        'time': 'Time:',
        'description': 'Description',
        'balance': 'Balance',
        'total': 'Total',
        'administrator': 'ADMINISTRATOR',
        'final_result': 'RESULT FOR THE PERIOD'
    }
}

# Dicionário de traduções dos grupos DRE (fallback quando não há tradução no banco)
GROUP_TRANSLATIONS = {
    'pt_to_en': {
        'RECEITA BRUTA': 'GROSS REVENUE',
        'RECEITAS': 'REVENUES',
        'VENDAS': 'SALES',
        'IMPOSTOS E DEDUÇÕES': 'TAXES AND DEDUCTIONS',
        'IMPOSTOS E DEDUÇOES': 'TAXES AND DEDUCTIONS',  # Variação sem acento
        'IMPOSTOS': 'TAXES',
        'DEDUÇÕES': 'DEDUCTIONS',
        'CUSTO DOS SERVIÇOS VENDIDOS - CSV': 'COST OF SERVICES SOLD - CSS',
        'CUSTO DOS SERVIÇOS VENDIDOS': 'COST OF SERVICES SOLD',
        'CUSTO DOS SERVICOS VENDIDOS - CSV': 'COST OF SERVICES SOLD - CSS',  # Variação sem acento
        'CUSTO DOS SERVICOS VENDIDOS': 'COST OF SERVICES SOLD',  # Variação sem acento
        'CUSTOS': 'COSTS',
        'CMV': 'COGS',
        'CSV': 'CSS',
        'DESPESAS ADMINISTRATIVAS': 'ADMINISTRATIVE EXPENSES',
        'DESPESAS COM VENDAS': 'SALES EXPENSES',
        'DESPESAS OPERACIONAIS': 'OPERATING EXPENSES',
        'DESPESAS': 'EXPENSES',
        'DEPRECIAÇÃO E AMORTIZAÇÃO': 'DEPRECIATION AND AMORTIZATION',
        'DEPRECIACAO E AMORTIZACAO': 'DEPRECIATION AND AMORTIZATION',  # Variação sem acento
        'DEPRECIAÇÃO': 'DEPRECIATION',
        'AMORTIZAÇÃO': 'AMORTIZATION',
        'RESULTADO FINANCEIRO': 'FINANCIAL RESULT',
        'RECEITAS FINANCEIRAS': 'FINANCIAL INCOME',
        'DESPESAS FINANCEIRAS': 'FINANCIAL EXPENSES',
        'IRPJ E CSLL': 'INCOME TAX AND CSLL',
        'IRPJ': 'INCOME TAX',
        'CSLL': 'CSLL',
        'IMPOSTOS SOBRE O LUCRO': 'INCOME TAXES',
        'PROVISÃO PARA IR E CSLL': 'PROVISION FOR INCOME TAX AND CSLL',
        'PROVISAO PARA IR E CSLL': 'PROVISION FOR INCOME TAX AND CSLL'  # Variação sem acento
    }
}

# ==============================================================================
# CONSULTAS SQL COM SUPORTE A INGLÊS
# ==============================================================================
def get_queries(codi_emp, data_inicial, data_final, ingles=False):
    # Determina se deve usar traduções em inglês
    if ingles:
        # JOINs para tradução das contas e grupos
        idioma_join_contas = """
            LEFT OUTER JOIN BETHADBA.CTCONTAS_IDIOMAS AS CTCONTAS_IDIOMAS 
            ON (CTCONTAS_IDIOMAS.CODI_EMP = CTCONTAS.CODI_EMP 
                AND CTCONTAS_IDIOMAS.CODI_CTA = CTCONTAS.CODI_CTA 
                AND CTCONTAS_IDIOMAS.I_IDIOMAS = 1)
        """
        idioma_join_grupos = """
            LEFT OUTER JOIN BETHADBA.CTGRUPOSDRE_IDIOMAS AS CTGRUPOSDRE_IDIOMAS 
            ON (CTGRUPOSDRE_IDIOMAS.CODI_EMP = CTGRUPOSDRE.CODI_EMP 
                AND CTGRUPOSDRE_IDIOMAS.CODIGO = CTGRUPOSDRE.CODIGO 
                AND CTGRUPOSDRE_IDIOMAS.I_IDIOMAS = 1)
        """
        # SELECT com fallback para português se não houver tradução
        idioma_select_conta = "COALESCE(CTCONTAS_IDIOMAS.DESCRICAO, CTCONTAS.NOME_CTA)"
        idioma_select_grupo = "COALESCE(CTGRUPOSDRE_IDIOMAS.DESCRICAO, CTGRUPOSDRE.DESCRICAO)"
    else:
        idioma_join_contas = ""
        idioma_join_grupos = ""
        idioma_select_conta = "CTCONTAS.NOME_CTA"
        idioma_select_grupo = "CTGRUPOSDRE.DESCRICAO"

    return {
        # DEIXAR ESPAÇO AQUI PARA COLAR A CONSULTA COMPLETA
        "estrutura_dre_completa_tipo2": f"""
            -- DRE NORMAL - CONTAS ANALÍTICAS
            SELECT 1 AS DEMONSTRATIVO, 
                   CTGRUPOSDRE.SEQUENCIA AS SEQUENCIA, 
                   DSDBA.FG_MONTA_MASCARA_CT('#.#.#.###.####', CTCONTAS.CLAS_CTA) AS CLASCTA, 
                   GEEMPRE.NOME_EMP AS NOMEEMP, 
                   TD_CONTADOR.NOME_CON AS NOMCONT, 
                   COALESCE(TD_CONTADOR.RCRC_CON, ' ') AS CRCCONT, 
                   GEEMPRE.RLEG_EMP AS RESPON, 
                   GEEMPRE.CARGO_LEG_EMP AS CARGO, 
                   DATE('{data_final}') AS DATAFIN, 
                   DATE('{data_inicial}') AS DATAINI, 
                   CAST({idioma_select_grupo} AS VARCHAR(90)) AS NOMEGRUPO, 
                   CAST({idioma_select_conta} AS VARCHAR(90)) AS NOMECONTA, 
                   0 AS SUBQUEBRA, 0 AS TEMTOTAL1, 0 AS TEMTOTAL2, 
                   CAST(0 AS DECIMAL(13,2)) AS VALOR, 
                   CAST(0 AS DECIMAL(13,2)) AS TOTAL1, 
                   CAST(0 AS DECIMAL(13,2)) AS TOTAL2, 
                   CTGRUPOSDRE.OPERACAO AS GRUPOTEMVALOR, 
                   CAST(0 AS DECIMAL(13,2)) AS VALORGRUPO, 
                   CTCONTAS.CODI_CTA AS CODICTA, 
                   GEEMPRE.ESTA_EMP AS ESTAEMP, 
                   GEEMPRE.CGCE_EMP AS CGCEEMP, 
                   CTGRUPOSDRE.CONFIGURACAO AS CTGRUPOSDRE_CONFIGURACAO, 
                   CTGRUPOSDRE.SOMA AS CTGRUPOSDRE_SOMA, 
                   CTGRUPOSDRE.GRUPOS AS CTGRUPOSDRE_GRUPOS, 
                   CTGRUPOSDRE.IMPRIMIR_GRUPO AS IMPRIMIR_GRUPO, 
                   'T' AS CTGRUPOSDRE_TIPO, 
                   CURRENT TIMESTAMP AS EMISSAO, 
                   GEEMPRE.IJUC_EMP AS INSC_JCOM, 
                   GEEMPRE.DJUC_EMP AS DATA_JCOM, 
                   (CASE WHEN TD_LUCRO.TEXTO = '' THEN 'LUCRO LÍQUIDO DO EXERCÍCIO' ELSE TD_LUCRO.TEXTO END) AS TEXTO_LUCRO, 
                   (CASE WHEN TD_PREJUIZO.TEXTO = '' THEN 'PREJUÍZO DO EXERCÍCIO' ELSE TD_PREJUIZO.TEXTO END) AS TEXTO_PREJUIZO, 
                   CTGRUPOSDRE.CODIGO AS CODIGO, 
                   TD_TITULO.TITULO AS TITULO_DRE, 
                   TD_TITULO_SEGUNDA_LINHA.TITULO AS TITULO_DRE_SEGUNDA_LINHA, 
                   '' AS TITULO_DRA, '' AS CODICTA_REF, '' AS CLASCTA_REF 
            FROM BETHADBA.GEEMPRE AS GEEMPRE 
                 LEFT OUTER JOIN(SELECT C.NOME_CON AS NOME_CON, C.RCRC_CON AS RCRC_CON, C.CODI_CON AS CODI_CON 
                                FROM BETHADBA.GECONTADOR AS C) AS TD_CONTADOR ON TD_CONTADOR.CODI_CON = GEEMPRE.CODI_CON, 
                 BETHADBA.CTCONTAS AS CTCONTAS {idioma_join_contas}, 
                 BETHADBA.CTGRUPOSDRE AS CTGRUPOSDRE {idioma_join_grupos}, 
                 LATERAL(SELECT COALESCE(TRIM(MAX(CAST(I.VALOR AS VARCHAR(5000)))), '') AS TEXTO 
                        FROM BETHADBA.GEINICIAL AS I 
                        WHERE I.CHAVE = 'DRE_descricao_resultado' 
                          AND I.SECAO = 'Lucro_' || CAST({codi_emp} AS CHAR(7))) AS TD_LUCRO, 
                 LATERAL(SELECT COALESCE(TRIM(MAX(CAST(I.VALOR AS VARCHAR(5000)))), '') AS TEXTO 
                        FROM BETHADBA.GEINICIAL AS I 
                        WHERE I.CHAVE = 'DRE_descricao_resultado' 
                          AND I.SECAO = 'Prejuizo_' || CAST({codi_emp} AS CHAR(7))) AS TD_PREJUIZO, 
                 LATERAL(SELECT COALESCE(TRIM(MAX(CAST(I.VALOR AS VARCHAR(5000)))), '') AS TITULO 
                        FROM BETHADBA.GEINICIAL AS I 
                        WHERE I.CHAVE = 'DRE' 
                          AND I.SECAO = 'titulo_' || CAST({codi_emp} AS CHAR(7))) AS TD_TITULO, 
                 LATERAL(SELECT COALESCE(TRIM(MAX(CAST(I.VALOR AS VARCHAR(5000)))), '') AS TITULO 
                        FROM BETHADBA.GEINICIAL AS I 
                        WHERE I.CHAVE = 'DRE' 
                          AND I.SECAO = 'titulo_linha2_' || CAST({codi_emp} AS CHAR(7))) AS TD_TITULO_SEGUNDA_LINHA 
            WHERE GEEMPRE.CODI_EMP = {codi_emp} 
              AND CTCONTAS.CODI_EMP = {codi_emp} 
              AND CTCONTAS.GRDRE_CTA IS NOT NULL 
              AND CTCONTAS.GRDRE_CTA > 0 
              AND CTGRUPOSDRE.CODI_EMP = {codi_emp} 
              AND CTCONTAS.GRDRE_CTA = CTGRUPOSDRE.CODIGO 
              AND CTCONTAS.TIPO_CTA = 'A'  -- Apenas contas analíticas

            UNION ALL

            -- DRE NORMAL - CONTAS SINTÉTICAS EXPANDIDAS
            SELECT 1 AS DEMONSTRATIVO, 
                   CTGRUPOSDRE.SEQUENCIA AS SEQUENCIA, 
                   DSDBA.FG_MONTA_MASCARA_CT('#.#.#.###.####', TDCONTA.CLAS_CTA) AS CLASCTA, 
                   GEEMPRE.NOME_EMP AS NOMEEMP, 
                   TD_CONTADOR.NOME_CON AS NOMCONT, 
                   COALESCE(TD_CONTADOR.RCRC_CON, ' ') AS CRCCONT, 
                   GEEMPRE.RLEG_EMP AS RESPON, 
                   GEEMPRE.CARGO_LEG_EMP AS CARGO, 
                   DATE('{data_final}') AS DATAFIN, 
                   DATE('{data_inicial}') AS DATAINI, 
                   CAST({idioma_select_grupo} AS VARCHAR(90)) AS NOMEGRUPO, 
                   CAST(TDCONTA.NOME_CTA AS VARCHAR(90)) AS NOMECONTA, 
                   0 AS SUBQUEBRA, 0 AS TEMTOTAL1, 0 AS TEMTOTAL2, 
                   CAST(0 AS DECIMAL(13,2)) AS VALOR, 
                   CAST(0 AS DECIMAL(13,2)) AS TOTAL1, 
                   CAST(0 AS DECIMAL(13,2)) AS TOTAL2, 
                   CTGRUPOSDRE.OPERACAO AS GRUPOTEMVALOR, 
                   CAST(0 AS DECIMAL(13,2)) AS VALORGRUPO, 
                   TDCONTA.CODI_CTA AS CODICTA, 
                   GEEMPRE.ESTA_EMP AS ESTAEMP, 
                   GEEMPRE.CGCE_EMP AS CGCEEMP, 
                   CTGRUPOSDRE.CONFIGURACAO AS CTGRUPOSDRE_CONFIGURACAO, 
                   CTGRUPOSDRE.SOMA AS CTGRUPOSDRE_SOMA, 
                   CTGRUPOSDRE.GRUPOS AS CTGRUPOSDRE_GRUPOS, 
                   CTGRUPOSDRE.IMPRIMIR_GRUPO AS IMPRIMIR_GRUPO, 
                   'T' AS CTGRUPOSDRE_TIPO, 
                   CURRENT TIMESTAMP AS EMISSAO, 
                   GEEMPRE.IJUC_EMP AS INSC_JCOM, 
                   GEEMPRE.DJUC_EMP AS DATA_JCOM, 
                   (CASE WHEN TD_LUCRO.TEXTO = '' THEN 'LUCRO LÍQUIDO DO EXERCÍCIO' ELSE TD_LUCRO.TEXTO END) AS TEXTO_LUCRO, 
                   (CASE WHEN TD_PREJUIZO.TEXTO = '' THEN 'PREJUÍZO DO EXERCÍCIO' ELSE TD_PREJUIZO.TEXTO END) AS TEXTO_PREJUIZO, 
                   CTGRUPOSDRE.CODIGO AS CODIGO, 
                   TD_TITULO.TITULO AS TITULO_DRE, 
                   TD_TITULO_SEGUNDA_LINHA.TITULO AS TITULO_DRE_SEGUNDA_LINHA, 
                   '' AS TITULO_DRA, '' AS CODICTA_REF, '' AS CLASCTA_REF 
            FROM BETHADBA.GEEMPRE AS GEEMPRE 
                 LEFT OUTER JOIN(SELECT C.NOME_CON AS NOME_CON, C.RCRC_CON AS RCRC_CON, C.CODI_CON AS CODI_CON 
                                FROM BETHADBA.GECONTADOR AS C) AS TD_CONTADOR ON TD_CONTADOR.CODI_CON = GEEMPRE.CODI_CON, 
                 BETHADBA.CTCONTAS AS CTCONTAS {idioma_join_contas}, 
                 BETHADBA.CTGRUPOSDRE AS CTGRUPOSDRE {idioma_join_grupos}, 
                 LATERAL(SELECT COALESCE(TRIM(MAX(CAST(I.VALOR AS VARCHAR(5000)))), '') AS TEXTO 
                        FROM BETHADBA.GEINICIAL AS I 
                        WHERE I.CHAVE = 'DRE_descricao_resultado' 
                          AND I.SECAO = 'Lucro_' || CAST({codi_emp} AS CHAR(7))) AS TD_LUCRO, 
                 LATERAL(SELECT COALESCE(TRIM(MAX(CAST(I.VALOR AS VARCHAR(5000)))), '') AS TEXTO 
                        FROM BETHADBA.GEINICIAL AS I 
                        WHERE I.CHAVE = 'DRE_descricao_resultado' 
                          AND I.SECAO = 'Prejuizo_' || CAST({codi_emp} AS CHAR(7))) AS TD_PREJUIZO, 
                 LATERAL(SELECT COALESCE(TRIM(MAX(CAST(I.VALOR AS VARCHAR(5000)))), '') AS TITULO 
                        FROM BETHADBA.GEINICIAL AS I 
                        WHERE I.CHAVE = 'DRE' 
                          AND I.SECAO = 'titulo_' || CAST({codi_emp} AS CHAR(7))) AS TD_TITULO, 
                 LATERAL(SELECT COALESCE(TRIM(MAX(CAST(I.VALOR AS VARCHAR(5000)))), '') AS TITULO 
                        FROM BETHADBA.GEINICIAL AS I 
                        WHERE I.CHAVE = 'DRE' 
                          AND I.SECAO = 'titulo_linha2_' || CAST({codi_emp} AS CHAR(7))) AS TD_TITULO_SEGUNDA_LINHA, 
                 LATERAL(SELECT C.CODI_CTA AS CODI_CTA, C.CLAS_CTA AS CLAS_CTA, C.NOME_CTA AS NOME_CTA 
                        FROM BETHADBA.CTCONTAS AS C 
                        WHERE C.CODI_EMP = {codi_emp} 
                          AND ((CTCONTAS.TIPO_CTA = 'A' AND C.CODI_CTA = CTCONTAS.CODI_CTA) 
                               OR (CTCONTAS.TIPO_CTA = 'S' 
                                   AND LEFT(C.CLAS_CTA, LENGTH(CTCONTAS.CLAS_CTA)) = CTCONTAS.CLAS_CTA 
                                   AND NOT EXISTS(SELECT 1 FROM BETHADBA.CTCONTAS AS CONTA 
                                                 WHERE CONTA.CODI_EMP = C.CODI_EMP 
                                                   AND CONTA.CODI_CTA = C.CODI_CTA 
                                                   AND CONTA.GRDRE_CTA = CTCONTAS.GRDRE_CTA 
                                                   AND CONTA.TIPO_CTA = 'A') 
                                   AND (NOT EXISTS(SELECT 1 FROM BETHADBA.CTCONTAS AS CONTA 
                                                  WHERE CONTA.CODI_EMP = CTCONTAS.CODI_EMP 
                                                    AND CONTA.TIPO_CTA = 'S' 
                                                    AND LEFT(CONTA.CLAS_CTA, LENGTH(CTCONTAS.CLAS_CTA)) = CTCONTAS.CLAS_CTA 
                                                    AND LENGTH(CONTA.CLAS_CTA) > LENGTH(CTCONTAS.CLAS_CTA)) 
                                        OR C.CODI_CTA = CTCONTAS.CODI_CTA)))) AS TDCONTA 
            WHERE GEEMPRE.CODI_EMP = {codi_emp} 
              AND CTCONTAS.CODI_EMP = {codi_emp} 
              AND CTCONTAS.GRDRE_CTA IS NOT NULL 
              AND CTCONTAS.GRDRE_CTA > 0 
              AND CTGRUPOSDRE.CODI_EMP = {codi_emp} 
              AND CTCONTAS.GRDRE_CTA = CTGRUPOSDRE.CODIGO 
              AND CTCONTAS.TIPO_CTA = 'S'  -- Contas sintéticas

            UNION ALL

            -- GRUPOS VAZIOS (sem contas vinculadas mas devem aparecer)
            SELECT 1 AS DEMONSTRATIVO, 
                   CTGRUPOSDRE.SEQUENCIA AS SEQUENCIA, 
                   CAST(SPACE(14) AS VARCHAR(14)) AS CLASCTA, 
                   GEEMPRE.NOME_EMP AS NOMEEMP, 
                   TD_CONTADOR.NOME_CON AS NOMCONT, 
                   COALESCE(TD_CONTADOR.RCRC_CON, ' ') AS CRCCONT, 
                   GEEMPRE.RLEG_EMP AS RESPON, 
                   GEEMPRE.CARGO_LEG_EMP AS CARGO, 
                   DATE('{data_final}') AS DATAFIN, 
                   DATE('{data_inicial}') AS DATAINI, 
                   CAST({idioma_select_grupo} AS VARCHAR(90)) AS NOMEGRUPO, 
                   CAST(SPACE(40) AS VARCHAR(40)) AS NOMECONTA, 
                   0 AS SUBQUEBRA, 0 AS TEMTOTAL1, 0 AS TEMTOTAL2, 
                   CAST(0 AS DECIMAL(13,2)) AS VALOR, 
                   CAST(0 AS DECIMAL(13,2)) AS TOTAL1, 
                   CAST(0 AS DECIMAL(13,2)) AS TOTAL2, 
                   CTGRUPOSDRE.OPERACAO AS GRUPOTEMVALOR, 
                   CAST(0 AS DECIMAL(13,2)) AS VALORGRUPO, 
                   0 AS CODICTA, 
                   GEEMPRE.ESTA_EMP AS ESTAEMP, 
                   GEEMPRE.CGCE_EMP AS CGCEEMP, 
                   CTGRUPOSDRE.CONFIGURACAO AS CTGRUPOSDRE_CONFIGURACAO, 
                   CTGRUPOSDRE.SOMA AS CTGRUPOSDRE_SOMA, 
                   CTGRUPOSDRE.GRUPOS AS CTGRUPOSDRE_GRUPOS, 
                   CTGRUPOSDRE.IMPRIMIR_GRUPO AS IMPRIMIR_GRUPO, 
                   'T' AS CTGRUPOSDRE_TIPO, 
                   CURRENT TIMESTAMP AS EMISSAO, 
                   GEEMPRE.IJUC_EMP AS INSC_JCOM, 
                   GEEMPRE.DJUC_EMP AS DATA_JCOM, 
                   (CASE WHEN TD_LUCRO.TEXTO = '' THEN 'LUCRO LÍQUIDO DO EXERCÍCIO' ELSE TD_LUCRO.TEXTO END) AS TEXTO_LUCRO, 
                   (CASE WHEN TD_PREJUIZO.TEXTO = '' THEN 'PREJUÍZO DO EXERCÍCIO' ELSE TD_PREJUIZO.TEXTO END) AS TEXTO_PREJUIZO, 
                   CTGRUPOSDRE.CODIGO AS CODIGO, 
                   TD_TITULO.TITULO AS TITULO_DRE, 
                   TD_TITULO_SEGUNDA_LINHA.TITULO AS TITULO_DRE_SEGUNDA_LINHA, 
                   '' AS TITULO_DRA, '' AS CODICTA_REF, '' AS CLASCTA_REF 
            FROM BETHADBA.GEEMPRE AS GEEMPRE 
                 LEFT OUTER JOIN(SELECT C.NOME_CON AS NOME_CON, C.RCRC_CON AS RCRC_CON, C.CODI_CON AS CODI_CON 
                                FROM BETHADBA.GECONTADOR AS C) AS TD_CONTADOR ON TD_CONTADOR.CODI_CON = GEEMPRE.CODI_CON, 
                 BETHADBA.CTGRUPOSDRE AS CTGRUPOSDRE {idioma_join_grupos}, 
                 LATERAL(SELECT COALESCE(TRIM(MAX(CAST(I.VALOR AS VARCHAR(5000)))), '') AS TEXTO 
                        FROM BETHADBA.GEINICIAL AS I 
                        WHERE I.CHAVE = 'DRE_descricao_resultado' 
                          AND I.SECAO = 'Lucro_' || CAST({codi_emp} AS CHAR(7))) AS TD_LUCRO, 
                 LATERAL(SELECT COALESCE(TRIM(MAX(CAST(I.VALOR AS VARCHAR(5000)))), '') AS TEXTO 
                        FROM BETHADBA.GEINICIAL AS I 
                        WHERE I.CHAVE = 'DRE_descricao_resultado' 
                          AND I.SECAO = 'Prejuizo_' || CAST({codi_emp} AS CHAR(7))) AS TD_PREJUIZO, 
                 LATERAL(SELECT COALESCE(TRIM(MAX(CAST(I.VALOR AS VARCHAR(5000)))), '') AS TITULO 
                        FROM BETHADBA.GEINICIAL AS I 
                        WHERE I.CHAVE = 'DRE' 
                          AND I.SECAO = 'titulo_' || CAST({codi_emp} AS CHAR(7))) AS TD_TITULO, 
                 LATERAL(SELECT COALESCE(TRIM(MAX(CAST(I.VALOR AS VARCHAR(5000)))), '') AS TITULO 
                        FROM BETHADBA.GEINICIAL AS I 
                        WHERE I.CHAVE = 'DRE' 
                          AND I.SECAO = 'titulo_linha2_' || CAST({codi_emp} AS CHAR(7))) AS TD_TITULO_SEGUNDA_LINHA 
            WHERE GEEMPRE.CODI_EMP = {codi_emp} 
              AND CTGRUPOSDRE.OPERACAO IN (2,3)  -- Grupos de totalização
              AND CTGRUPOSDRE.CODI_EMP = {codi_emp}
              
            ORDER BY 1, 2, 39, 3, 12
        """,
        
        "cabecalho_empresa": f"""
            SELECT
                GEEMPRE.NOME_EMP,
                GEEMPRE.CGCE_EMP,
                CTPARMTO.DINR_PAR,
                CTPARMTO.DFIR_PAR,
                CTPARMTO.NLIV_PAR,
                CTPARMTO.FINI_PAR
            FROM BETHADBA.GEEMPRE
            JOIN BETHADBA.CTPARMTO ON GEEMPRE.CODI_EMP = CTPARMTO.CODI_EMP
            WHERE GEEMPRE.CODI_EMP = {codi_emp}
        """,
        
        "cabecalho_contador": f"""
            SELECT NOME_CON, RCRC_CON, CPFC_CON FROM BETHADBA.GECONTADOR WHERE GECONTADOR.CODI_CON = 5
        """,
        
        "cabecalho_administrador": f"""
            SELECT RLEG_EMP, CPF_LEG_EMP FROM BETHADBA.GEEMPRE WHERE CODI_EMP = {codi_emp}
        """,
        
        "saldos_contas": f"""
            SELECT 
                CTCONTAS.CODI_CTA AS CODI_CTA, 
                (COALESCE(TDTOTAL_DEBITO.VALOR_TOTAL, 0) - COALESCE(TDTOTAL_CREDITO.VALOR_TOTAL, 0)) AS SALDOATU
            FROM BETHADBA.CTCONTAS AS CTCONTAS, 
                 BETHADBA.CTPARMTO AS CTPARMTO, 
                 LATERAL(SELECT COALESCE(SUM(LANX.VLOR_LAN), 0) AS VALOR_TOTAL 
                        FROM BETHADBA.CTLANCTO LANX 
                        WHERE LANX.CODI_EMP = {codi_emp} 
                          AND LANX.DATA_LAN >= '{data_inicial}' 
                          AND LANX.DATA_LAN <= '{data_final}' 
                          AND (1 = 0 OR LANX.ORIG_LAN <> 2) 
                          AND LANX.CODI_EMP_PLANO = CTCONTAS.CODI_EMP 
                          AND LANX.CDEB_LAN = CTCONTAS.CODI_CTA) TDTOTAL_DEBITO,
                 LATERAL(SELECT COALESCE(SUM(LANX.VLOR_LAN), 0) AS VALOR_TOTAL 
                        FROM BETHADBA.CTLANCTO LANX 
                        WHERE LANX.CODI_EMP = {codi_emp} 
                          AND LANX.DATA_LAN >= '{data_inicial}' 
                          AND LANX.DATA_LAN <= '{data_final}' 
                          AND (1 = 0 OR LANX.ORIG_LAN <> 2) 
                          AND LANX.CODI_EMP_PLANO = CTCONTAS.CODI_EMP 
                          AND LANX.CCRE_LAN = CTCONTAS.CODI_CTA) TDTOTAL_CREDITO 
            WHERE CTCONTAS.CODI_EMP = {codi_emp} 
              AND CTPARMTO.CODI_EMP = {codi_emp}
        """
    }

# --- 2. FUNÇÕES AUXILIARES E DE LÓGICA ---

def fetch_data(conn_str, queries):
    results = {}
    try:
        with pyodbc.connect(conn_str) as conn:
            print("Conexão com o banco de dados bem-sucedida.")
            with conn.cursor() as cursor:
                for key, query in queries.items():
                    print(f"Executando consulta: {key}...")
                    cursor.execute(query)
                    columns = [column[0].upper() for column in cursor.description]
                    rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
                    results[key] = rows
                    print(f"Consulta {key} retornou {len(rows)} linhas.")
    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"ERRO de Conexão ou Consulta: {sqlstate}")
        print(ex)
        return None
    print("Conexão fechada.")
    return results

def salvar_dados_json(data, codi_emp, data_inicial, data_final, ingles=False, timestamp=None):
    def convert_types(obj):
        if hasattr(obj, 'isoformat'):
            return obj.isoformat()
        elif isinstance(obj, Decimal):
            return float(obj)
        raise TypeError(f"Object of type {type(obj)} is not JSON serializable")
    
    data_ini_fmt = data_inicial.replace('-', '')
    data_fim_fmt = data_final.replace('-', '')
    lang_suffix = "_EN" if ingles else ""
    timestamp_suffix = f"_{timestamp}" if timestamp else ""
    json_filename = f"dados_dre_EMP{codi_emp}_{data_ini_fmt}_a_{data_fim_fmt}{lang_suffix}{timestamp_suffix}.json"
    
    with open(json_filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=convert_types)
    
    print(f"Dados salvos em '{json_filename}'")
    return json_filename

def format_currency(value):
    if value is None:
        value = Decimal('0.00')
    if not isinstance(value, Decimal):
        value = Decimal(str(value))
    
    value = value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    
    formatted_value = f"{abs(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    
    if value < 0:
        return f"({formatted_value})"
        
    return formatted_value

# --- 3. GERADOR DRE FLEXÍVEL COM CONSULTA ORIGINAL ---

class DREGenerator:
    def __init__(self, data, codi_emp, data_inicial, data_final, ingles=False, timestamp=None):
        self.data = data
        self.codi_emp = codi_emp
        self.data_inicial = data_inicial
        self.data_final = data_final
        self.ingles = ingles
        self.timestamp = timestamp
        self.lang = 'en' if ingles else 'pt'
        self.t = TRANSLATIONS[self.lang]  # Atalho para traduções
        self.story = []
        self.styles = getSampleStyleSheet()
        self.doc = None
        self.current_datetime = datetime.now()
        self.header_info = self._extract_header_info()
        self.tahoma_available = self._register_tahoma_font()
        self._setup_styles()

    def _register_tahoma_font(self):
        try:
            pdfmetrics.registerFont(TTFont('Tahoma', 'tahoma.ttf'))
            pdfmetrics.registerFont(TTFont('Tahoma-Bold', 'tahomab.ttf'))
            print("Fonte Tahoma registrada com sucesso")
            return True
        except Exception as e:
            print(f"Erro ao registrar Tahoma: {e}. Usando Helvetica como substituto.")
            return False
            
    def _setup_styles(self):
        font_name = 'Tahoma' if self.tahoma_available else 'Helvetica'
        font_bold = 'Tahoma-Bold' if self.tahoma_available else 'Helvetica-Bold'
        
        self.style_normal = ParagraphStyle(
            name='DataNormal', parent=self.styles['Normal'], fontName=font_name, 
            fontSize=7.6, leading=9.5, alignment=2 # RIGHT
        )
        self.style_normal_left = ParagraphStyle(
            name='DataNormalLeft', parent=self.style_normal, alignment=0 # LEFT
        )
        self.style_bold_left = ParagraphStyle(
            name='DataBoldLeft', parent=self.style_normal_left, fontName=font_bold
        )
        self.style_bold_right = ParagraphStyle(
            name='DataBoldRight', parent=self.style_normal, fontName=font_bold
        )

    def _extract_header_info(self):
        info = self.data['cabecalho_empresa'][0]
        contador = self.data['cabecalho_contador'][0]
        admin = self.data['cabecalho_administrador'][0]
        
        cnpj = info['CGCE_EMP'].strip()
        cnpj_fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
        
        # CORREÇÃO: Usar as datas dos parâmetros ao invés da tabela
        start_date = datetime.strptime(self.data_inicial, '%Y-%m-%d').strftime('%d/%m/%Y')
        end_date = datetime.strptime(self.data_final, '%Y-%m-%d').strftime('%d/%m/%Y')

        return {
            "empresa": info['NOME_EMP'].strip(), 
            "cnpj": cnpj_fmt, 
            "periodo": f"{start_date} - {end_date}",  # ← Agora usa os parâmetros corretos
            "data_emissao": self.current_datetime.strftime('%d/%m/%Y'), 
            "hora_emissao": self.current_datetime.strftime('%H:%M:%S'),
            "folha": f"{info['FINI_PAR']:04d}", 
            "livro": f"{info['NLIV_PAR']:04d}",
            "contador_nome": contador['NOME_CON'].strip(), 
            "contador_crc": contador['RCRC_CON'].strip(),
            "admin_nome": admin['RLEG_EMP'].strip(), 
            "admin_cpf": admin['CPF_LEG_EMP'].strip(),
        }

    def _header_footer(self, canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.drawString(1.5 * cm, A4[1] - 1.5 * cm, f"{self.t['company']}")
        canvas.drawString(1.5 * cm, A4[1] - 2.0 * cm, f"{self.t['cnpj']}")
        canvas.drawString(1.5 * cm, A4[1] - 2.5 * cm, f"{self.t['period']}")
        
        canvas.setFont('Helvetica-Bold', 8)
        canvas.drawString(3.5 * cm, A4[1] - 1.5 * cm, self.header_info['empresa'])
        canvas.setFont('Helvetica', 8)
        canvas.drawString(3.5 * cm, A4[1] - 2.0 * cm, self.header_info['cnpj'])
        canvas.drawString(3.5 * cm, A4[1] - 2.5 * cm, self.header_info['periodo'])
        
        canvas.setFont('Helvetica-Bold', 10)
        report_date = datetime.strptime(self.header_info['periodo'].split(' - ')[1], '%d/%m/%Y').strftime('%d/%m/%Y')
        if self.ingles:
            title_text = f"{self.t['report_title']} {report_date}"  # P&L 28/02/2025
        else:
            title_text = f"{self.t['report_title']} EM {report_date}"  # DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO EM 28/02/2025
        canvas.drawCentredString(A4[0] / 2, A4[1] - 4.0 * cm, title_text)

        canvas.setFont('Helvetica', 8)
        canvas.drawString(A4[0] - 5 * cm, A4[1] - 1.5 * cm, f"{self.t['page']}")
        canvas.drawString(A4[0] - 5 * cm, A4[1] - 2.0 * cm, f"{self.t['book_number']}")
        canvas.drawString(A4[0] - 5 * cm, A4[1] - 2.5 * cm, f"{self.t['emission']}")
        canvas.drawString(A4[0] - 5 * cm, A4[1] - 3.0 * cm, f"{self.t['time']}")

        canvas.drawString(A4[0] - 3 * cm, A4[1] - 1.5 * cm, self.header_info['folha'])
        canvas.drawString(A4[0] - 3 * cm, A4[1] - 2.0 * cm, self.header_info['livro'])
        canvas.drawString(A4[0] - 3 * cm, A4[1] - 2.5 * cm, self.header_info['data_emissao'])
        canvas.drawString(A4[0] - 3 * cm, A4[1] - 3.0 * cm, self.header_info['hora_emissao'])
        canvas.restoreState()
    
    def _translate_group_name(self, group_name):
        """Traduz nomes de grupos para inglês usando fallback manual"""
        if not self.ingles:
            return group_name
        
        # Primeiro tenta tradução exata
        translated = GROUP_TRANSLATIONS['pt_to_en'].get(group_name.upper(), None)
        if translated:
            return translated
        
        # Se não encontrar tradução exata, tenta palavras-chave
        group_upper = group_name.upper()
        
        # Lógica de fallback por palavras-chave
        if 'RECEITA' in group_upper and 'BRUTA' in group_upper:
            return 'GROSS REVENUE'
        elif 'RECEITA' in group_upper:
            return 'REVENUE'
        elif 'IMPOSTOS' in group_upper and 'DEDUÇ' in group_upper:
            return 'TAXES AND DEDUCTIONS'
        elif 'IMPOSTOS' in group_upper:
            return 'TAXES'
        elif 'CUSTO' in group_upper and ('SERVIÇOS' in group_upper or 'SERVICOS' in group_upper):
            return 'COST OF SERVICES SOLD'
        elif 'CUSTO' in group_upper:
            return 'COSTS'
        elif 'DESPESAS' in group_upper and 'ADMINISTRATIVA' in group_upper:
            return 'ADMINISTRATIVE EXPENSES'
        elif 'DESPESAS' in group_upper and 'VENDAS' in group_upper:
            return 'SALES EXPENSES'
        elif 'DESPESAS' in group_upper:
            return 'EXPENSES'
        elif 'DEPRECIAÇÃO' in group_upper or 'DEPRECIACAO' in group_upper:
            return 'DEPRECIATION AND AMORTIZATION'
        elif 'RESULTADO' in group_upper and 'FINANCEIRO' in group_upper:
            return 'FINANCIAL RESULT'
        elif 'IRPJ' in group_upper and 'CSLL' in group_upper:
            return 'INCOME TAX AND CSLL'
        elif 'IRPJ' in group_upper:
            return 'INCOME TAX'
        
        # Se não encontrou nenhuma tradução, retorna o original
        return group_name

    def _create_assinaturas(self):
        style = ParagraphStyle(
            name='Signature', parent=self.styles['Normal'], fontName=self.style_normal.fontName, 
            fontSize=6.4, leading=8, alignment=1  # CENTER alignment
        )
        
        admin_cpf = self.header_info['admin_cpf']
        admin_cpf_fmt = f"{admin_cpf[:3]}.{admin_cpf[3:6]}.{admin_cpf[6:9]}-{admin_cpf[9:]}"
        
        if self.ingles:
            contador_crc_fmt = f"Registered in CRC - SP under No. {self.header_info['contador_crc']}"
            cpf_label = "SSN:"  # ← CPF pessoal = SSN em inglês
        else:
            contador_crc_fmt = f"Reg. no CRC - SP sob o No. {self.header_info['contador_crc']}"
            cpf_label = "CPF:"

        admin_block = [
            Paragraph("_______________________________________", style), 
            Spacer(1, 4), 
            Paragraph(self.header_info['admin_nome'], style), 
            Paragraph(self.t['administrator'], style), 
            Paragraph(f"{cpf_label} {admin_cpf_fmt}", style)
        ]
        contador_block = [
            Paragraph("_______________________________________", style), 
            Spacer(1, 4), 
            Paragraph(self.header_info['contador_nome'], style), 
            Paragraph(contador_crc_fmt, style)
        ]
        
        return admin_block, contador_block

    def _add_group_to_table(self, table_data, excel_data, group_name, group_total, is_calculated_subtotal=False):
        """Adiciona APENAS o grupo à tabela (sem contas individuais)"""
        # Traduzir nome do grupo se necessário
        translated_name = self._translate_group_name(group_name)
        
        # REGRA CORRIGIDA:
        # Grupos do banco (dados reais): aparecem em Saldo E Total
        # Subtotais calculados: aparecem apenas em Total
        
        if is_calculated_subtotal:
            # Subtotais calculados - apenas na coluna Total
            saldo_value = ''
            total_value = format_currency(group_total)
            excel_saldo = None
            excel_total = group_total
        else:
            # Grupos reais do banco - nas duas colunas
            saldo_value = format_currency(group_total)
            total_value = format_currency(group_total)
            excel_saldo = group_total
            excel_total = group_total
        
        table_data.append([
            Paragraph(f"<b>{translated_name}</b>", self.style_bold_left),
            Paragraph(saldo_value, self.style_bold_right) if saldo_value else '', 
            Paragraph(total_value, self.style_bold_right)
        ])
        excel_data.append({'type': 'group', 'values': [translated_name, excel_saldo, excel_total]})
        
        # Linha em branco após cada grupo
        table_data.append(['', '', ''])
        excel_data.append({'type': 'spacer', 'values': ['', '', '']})

    def prepare_table_data(self):
        """VERSÃO CORRIGIDA - Ordem correta dos grupos"""
        
        # 1. Obter saldos das contas
        saldos = {item['CODI_CTA']: Decimal(str(item['SALDOATU'])) for item in self.data['saldos_contas']}
        
        # 2. Usar consulta original complexa
        estrutura = self.data.get('estrutura_dre_completa_tipo2', [])
        
        # 3. Agrupar por NOMEGRUPO e somar saldos
        group_totals = {}
        group_order = []

        for row in estrutura:
            group_name = row['NOMEGRUPO'].strip()
            codicta = row['CODICTA']
            saldo = saldos.get(codicta, Decimal('0.0'))

            if saldo.is_zero(): 
                continue
                
            saldo *= -1  # Inversão de sinal

            if group_name not in group_order: 
                group_order.append(group_name)
                
            group_totals[group_name] = group_totals.get(group_name, Decimal('0.0')) + saldo
        
        # 4. Filtrar grupos com saldo
        final_groups = {name: total for name, total in group_totals.items() if not total.is_zero()}

        # 5. Debug - mostrar grupos encontrados
        print(f"Grupos encontrados na DRE ({len(final_groups)} grupos):")
        for i, (grupo, total) in enumerate(final_groups.items(), 1):
            print(f"   {i}. '{grupo}' -> {format_currency(total)}")

        # 6. Montar cabeçalho da tabela
        pdf_data = [[
            Paragraph(f"<b>{self.t['description']}</b>", self.style_bold_left), 
            Paragraph(f"<b>{self.t['balance']}</b>", self.style_bold_right), 
            Paragraph(f"<b>{self.t['total']}</b>", self.style_bold_right)
        ]]
        
        excel_data = [{
            'type': 'header', 
            'values': [self.t['description'], self.t['balance'], self.t['total']]
        }]
        
        spacer_row = {'type': 'spacer', 'values': ['', '', '']}
        
        # 7. === PROCESSAMENTO NA ORDEM CORRETA ===
        
        # 1. RECEITA BRUTA (grupo real do banco)
        receita_bruta = Decimal('0.0')
        for grupo_name, total in final_groups.items():
            if self._is_revenue_group(grupo_name):
                receita_bruta = total
                self._add_group_to_table(pdf_data, excel_data, grupo_name, receita_bruta, is_calculated_subtotal=False)
                break

        # 2. IMPOSTOS E DEDUÇÕES (grupo real do banco)
        impostos = Decimal('0.0')
        for grupo_name, total in final_groups.items():
            if self._is_tax_group(grupo_name):
                impostos = total
                self._add_group_to_table(pdf_data, excel_data, grupo_name, impostos, is_calculated_subtotal=False)
                break

        # SUBTOTAL: RECEITA LÍQUIDA
        receita_liquida = receita_bruta + impostos
        net_revenue_text = "NET REVENUE" if self.ingles else "RECEITA LÍQUIDA"
        pdf_data.extend([[
            Paragraph(f"<b>{net_revenue_text}</b>", self.style_bold_left), 
            '', 
            Paragraph(f"<b>{format_currency(receita_liquida)}</b>", self.style_bold_right)
        ], ['','','']])
        excel_data.extend([{'type': 'total', 'values': [net_revenue_text, None, receita_liquida]}, spacer_row])

        # 3. CUSTOS CMV/CSV (grupos reais do banco)
        custo_total = Decimal('0.0')
        for grupo_name, total in final_groups.items():
            if self._is_cost_group(grupo_name):
                custo_total += total
                self._add_group_to_table(pdf_data, excel_data, grupo_name, total, is_calculated_subtotal=False)

        # SUBTOTAL: LUCRO BRUTO
        lucro_bruto = receita_liquida + custo_total
        gross_profit_text = "GROSS PROFIT" if self.ingles else "LUCRO BRUTO"
        pdf_data.extend([[
            Paragraph(f"<b>{gross_profit_text}</b>", self.style_bold_left), 
            '', 
            Paragraph(f"<b>{format_currency(lucro_bruto)}</b>", self.style_bold_right)
        ], ['','','']])
        excel_data.extend([{'type': 'total', 'values': [gross_profit_text, None, lucro_bruto]}, spacer_row])

        # 4. Calcular total das DESPESAS OPERACIONAIS (somar administrativas + vendas)
        despesas_administrativas = Decimal('0.0')
        despesas_vendas = Decimal('0.0')
        
        for grupo_name, total in final_groups.items():
            if self._is_administrative_expense_group(grupo_name):
                despesas_administrativas += total
            elif self._is_sales_expense_group(grupo_name):
                despesas_vendas += total

        despesas_operacionais_total = despesas_administrativas + despesas_vendas

        # 5. NOVO GRUPO: DESPESAS OPERACIONAIS (subtotal calculado) - ANTES das específicas
        operational_expenses_text = "OPERATING EXPENSES" if self.ingles else "DESPESAS OPERACIONAIS"
        self._add_group_to_table(pdf_data, excel_data, operational_expenses_text, despesas_operacionais_total, is_calculated_subtotal=True)

        # 6. DESPESAS ADMINISTRATIVAS (grupos reais do banco) - APÓS as operacionais
        for grupo_name, total in final_groups.items():
            if self._is_administrative_expense_group(grupo_name):
                self._add_group_to_table(pdf_data, excel_data, grupo_name, total, is_calculated_subtotal=False)

        # 7. DESPESAS COM VENDAS (grupos reais do banco) - APÓS as administrativas
        for grupo_name, total in final_groups.items():
            if self._is_sales_expense_group(grupo_name):
                self._add_group_to_table(pdf_data, excel_data, grupo_name, total, is_calculated_subtotal=False)

        # SUBTOTAL: RESULTADO OPERACIONAL (EBITDA)
        resultado_ebitda = lucro_bruto + despesas_operacionais_total
        operating_result_text = "OPERATING RESULT (EBITDA)" if self.ingles else "RESULTADO OPERACIONAL (EBITDA)"
        pdf_data.extend([[
            Paragraph(f"<b>{operating_result_text}</b>", self.style_bold_left), 
            '', 
            Paragraph(f"<b>{format_currency(resultado_ebitda)}</b>", self.style_bold_right)
        ], ['','','']])
        excel_data.extend([{'type': 'total', 'values': [operating_result_text, None, resultado_ebitda]}, spacer_row])
        
        # 8. DEPRECIAÇÃO E AMORTIZAÇÃO
        deprec_amort = Decimal('0.0')
        for grupo_name, total in final_groups.items():
            if self._is_depreciation_group(grupo_name):
                deprec_amort += total
                self._add_group_to_table(pdf_data, excel_data, grupo_name, total, is_calculated_subtotal=False)
        
        # 9. RESULTADO FINANCEIRO
        resultado_financeiro = Decimal('0.0')
        for grupo_name, total in final_groups.items():
            if self._is_financial_group(grupo_name):
                resultado_financeiro += total
                self._add_group_to_table(pdf_data, excel_data, grupo_name, total, is_calculated_subtotal=False)

        # SUBTOTAL: LUCRO ANTES DOS IMPOSTOS
        lucro_antes_impostos = resultado_ebitda + deprec_amort + resultado_financeiro
        profit_before_taxes_text = "PROFIT BEFORE TAXES" if self.ingles else "LUCRO ANTES DOS IMPOSTOS"
        pdf_data.extend([[
            Paragraph(f"<b>{profit_before_taxes_text}</b>", self.style_bold_left), 
            '', 
            Paragraph(f"<b>{format_currency(lucro_antes_impostos)}</b>", self.style_bold_right)
        ], ['','','']])
        excel_data.extend([{'type': 'total', 'values': [profit_before_taxes_text, None, lucro_antes_impostos]}, spacer_row])

        # 10. IMPOSTOS SOBRE O LUCRO
        impostos_lucro = Decimal('0.0')
        for grupo_name, total in final_groups.items():
            if self._is_income_tax_group(grupo_name):
                impostos_lucro += total
                self._add_group_to_table(pdf_data, excel_data, grupo_name, total, is_calculated_subtotal=False)

        # 11. Resultado final
        resultado_final = lucro_antes_impostos + impostos_lucro
        
        return pdf_data, excel_data, resultado_final
    
    def _is_revenue_group(self, group_name):
        """Identifica grupos de receita (português ou inglês)"""
        keywords_pt = ["RECEITA BRUTA", "RECEITA", "VENDAS", "VENDA"]
        keywords_en = ["GROSS REVENUE", "REVENUE", "SALES"]
        exclude = ["DEDUÇ", "TAX", "DEDUCTION"]
        
        group_upper = group_name.upper()
        has_revenue = (any(k in group_upper for k in keywords_pt) or 
                    any(k in group_upper for k in keywords_en))
        has_exclude = any(k in group_upper for k in exclude)
        
        return has_revenue and not has_exclude

    def _is_tax_group(self, group_name):
        """Identifica grupos de impostos e deduções"""
        keywords_pt = ["IMPOSTOS E DEDUÇÕES", "IMPOSTOS", "DEDUÇÕES"]
        keywords_en = ["TAX AND RETURNS", "TAXES", "DEDUCTIONS"]
        
        group_upper = group_name.upper()
        return (any(k in group_upper for k in keywords_pt) or 
                any(k in group_upper for k in keywords_en))

    def _is_cost_group(self, group_name):
        """Identifica grupos de custos (APENAS custos de produção/vendas)"""
        cost_keywords_pt = ["CUSTO DOS SERVIÇOS VENDIDOS", "CUSTO", "CUSTOS", "CMV", "CSV"]
        cost_keywords_en = ["COST OF SERVICES SOLD", "COST", "COSTS", "COGS", "CSS"]
        admin_keywords = ["ADMINISTRATIVE", "ADMINISTRATIVA", "DESPESAS", "EXPENSES", "ADM", "AE"]
        
        group_upper = group_name.upper()
        has_cost = (any(k in group_upper for k in cost_keywords_pt) or 
                    any(k in group_upper for k in cost_keywords_en))
        has_admin = any(k in group_upper for k in admin_keywords)
        
        return has_cost and not has_admin

    def _is_administrative_expense_group(self, group_name):
        """Identifica APENAS despesas administrativas (não vendas)"""
        keywords_pt = ["DESPESAS ADMINISTRATIVAS", "ADMINISTRATIVA"]
        keywords_en = ["ADMINISTRATIVE EXPENSES", "ADMINISTRATIVE"]
        exclude_keywords = ["VENDAS", "SALES", "DEPRECIA", "AMORTIZ", "FINANCEIRO", "FINANCIAL"]
        
        group_upper = group_name.upper()
        has_admin = (any(k in group_upper for k in keywords_pt) or 
                     any(k in group_upper for k in keywords_en))
        has_exclude = any(k in group_upper for k in exclude_keywords)
        
        return has_admin and not has_exclude

    def _is_sales_expense_group(self, group_name):
        """Identifica APENAS despesas com vendas"""
        keywords_pt = ["DESPESAS COM VENDAS", "VENDAS"]
        keywords_en = ["SALES EXPENSES", "SALES"]
        exclude_keywords = ["ADMINISTRATIVA", "ADMINISTRATIVE", "DEPRECIA", "AMORTIZ", "FINANCEIRO", "FINANCIAL"]
        
        group_upper = group_name.upper()
        has_sales = (any(k in group_upper for k in keywords_pt) or 
                     any(k in group_upper for k in keywords_en))
        has_exclude = any(k in group_upper for k in exclude_keywords)
        
        return has_sales and not has_exclude

    def _is_operational_expense_group(self, group_name):
        """Identifica grupos de despesas operacionais (administrativas OU vendas)"""
        return self._is_administrative_expense_group(group_name) or self._is_sales_expense_group(group_name)

    def _is_depreciation_group(self, group_name):
        """Identifica grupos de depreciação e amortização"""
        keywords_pt = ["DEPRECIAÇÃO", "AMORTIZAÇÃO"]
        keywords_en = ["DEPRECIATION", "AMORTIZATION"]
        
        group_upper = group_name.upper()
        return (any(k in group_upper for k in keywords_pt) or 
                any(k in group_upper for k in keywords_en))

    def _is_financial_group(self, group_name):
        """Identifica grupos de resultado financeiro"""
        keywords_pt = ["RESULTADO FINANCEIRO", "FINANCEIRO"]
        keywords_en = ["FINANCIAL RESULT", "FINANCIAL", "RESULTS OF FINANCIAL OPERATIONS"]
        
        group_upper = group_name.upper()
        return (any(k in group_upper for k in keywords_pt) or 
                any(k in group_upper for k in keywords_en))

    def _is_income_tax_group(self, group_name):
        """Identifica impostos sobre o lucro"""
        keywords_pt = ["IRPJ", "CSLL", "IMPOSTOS SOBRE"]
        keywords_en = ["INCOME TAX", "PROFIT TAX"]
        
        group_upper = group_name.upper()
        return (any(k in group_upper for k in keywords_pt) or 
                any(k in group_upper for k in keywords_en))

    def _generate_xlsx(self, excel_data, resultado_final, filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DRE" if not self.ingles else "Income Statement"
        ws.sheet_view.showGridLines = False

        font_bold = Font(name='Calibri', size=11, bold=True)
        font_small = Font(name='Calibri', size=9)
        
        # Formato contábil: positivos com espaço à direita para alinhar com negativos, negativos entre parênteses.
        number_format_accounting = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

        # Linha 1: Nome da empresa (esquerda) e Data de emissão (direita)
        ws['A1'] = self.header_info['empresa']
        ws['A1'].font = font_bold
        ws['C1'] = f"{self.t['emission']} {self.header_info['data_emissao']}"
        ws['C1'].font = Font(name='Calibri', size=9)
        ws['C1'].alignment = Alignment(horizontal='right')
        
        # Linha 2: CNPJ (esquerda) e Hora (direita)
        ws['A2'] = f"{self.t['cnpj']} {self.header_info['cnpj']}"
        ws['C2'] = f"{self.t['time']} {self.header_info['hora_emissao']}"
        ws['C2'].font = Font(name='Calibri', size=9)
        ws['C2'].alignment = Alignment(horizontal='right')
        
        # Linha 3: Período
        ws.merge_cells('A3:C3'); ws['A3'] = f"{self.t['period']} {self.header_info['periodo']}"
        
        # Título do relatório na linha 5
        report_date = self.header_info['periodo'].split(' - ')[1]
        if self.ingles:
            title_text = f"{self.t['report_title']} {report_date}"  # P&L 28/02/2025
        else:
            title_text = f"{self.t['report_title']} EM {report_date}"  # DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO EM 28/02/2025
        ws.merge_cells('A5:C5'); ws['A5'] = title_text; ws['A5'].font = font_bold; ws['A5'].alignment = Alignment(horizontal='center')
        
        current_row = 7

        for row_info in excel_data:
            row_type = row_info['type']
            desc, saldo, total = row_info['values']

            if row_type == 'spacer' and not any(row_info['values']):
                current_row += 1
                continue

            cell_A = ws.cell(row=current_row, column=1, value=desc)
            cell_B = ws.cell(row=current_row, column=2, value=saldo if saldo is not None else '')
            cell_C = ws.cell(row=current_row, column=3, value=total if total is not None else '')

            cell_B.number_format = number_format_accounting
            cell_C.number_format = number_format_accounting
            
            # Aplicar negrito para cabeçalhos, totais e grupos
            if row_type in ['header', 'total', 'group']:
                cell_A.font = font_bold
                if row_type == 'header':
                    cell_B.font = font_bold  # Cabeçalho "Saldo" em negrito
                cell_C.font = font_bold
            
            # Valores da coluna Saldo em negrito quando preenchidos
            if saldo is not None and saldo != '':
                cell_B.font = font_bold
            
            current_row += 1

        # Resultado final
        cell_A = ws.cell(row=current_row, column=1, value=self.t['final_result'])
        cell_C = ws.cell(row=current_row, column=3, value=resultado_final)
        
        cell_A.font = font_bold; cell_C.font = font_bold
        cell_C.number_format = number_format_accounting
        
        thin_top_border = Border(top=Side(style='thin'))
        double_bottom_border = Border(bottom=Side(style='double'))
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{current_row-1}'].border = thin_top_border
            ws[f'{col}{current_row}'].border = double_bottom_border

        # Pular algumas linhas antes das assinaturas
        current_row += 4

        # === ADICIONAR ASSINATURAS ===
        admin_cpf = self.header_info['admin_cpf']
        admin_cpf_fmt = f"{admin_cpf[:3]}.{admin_cpf[3:6]}.{admin_cpf[6:9]}-{admin_cpf[9:]}"
        
        if self.ingles:
            contador_crc_fmt = f"Registered in CRC - SP under No. {self.header_info['contador_crc']}"
            cpf_label = "SSN:"
        else:
            contador_crc_fmt = f"Reg. no CRC - SP sob o No. {self.header_info['contador_crc']}"
            cpf_label = "CPF:"

        # Linha de assinatura do Administrador e Contador (centralizadas)
        ws[f'A{current_row}'] = "____________________________________"
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'C{current_row}'] = "____________________________________"  
        ws[f'C{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 2
        
        # Nome do Administrador e Contador (centralizados)
        ws[f'A{current_row}'] = self.header_info['admin_nome']
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'C{current_row}'] = self.header_info['contador_nome']
        ws[f'C{current_row}'].font = font_small
        ws[f'C{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Cargo do Administrador e CRC do Contador (centralizados)
        ws[f'A{current_row}'] = self.t['administrator']
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'C{current_row}'] = contador_crc_fmt
        ws[f'C{current_row}'].font = font_small
        ws[f'C{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        
        # CPF do Administrador (centralizado)
        ws[f'A{current_row}'] = f"{cpf_label} {admin_cpf_fmt}"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 70
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        
        wb.save(filename)
        print(f"XLSX '{filename}' gerado com sucesso.")

    def run(self):
        data_ini_fmt = self.data_inicial.replace('-', '')
        data_fim_fmt = self.data_final.replace('-', '')
        lang_suffix = "_EN" if self.ingles else ""
        timestamp_suffix = f"_{self.timestamp}" if self.timestamp else ""
        base_filename = f"DRE_EMP{self.codi_emp}_{data_ini_fmt}_a_{data_fim_fmt}{lang_suffix}{timestamp_suffix}"
        
        pdf_data, excel_data, resultado_final = self.prepare_table_data()

        pdf_filename = f"{base_filename}.pdf"
        xlsx_filename = f"{base_filename}.xlsx"
        
        self._generate_pdf(pdf_data, resultado_final, pdf_filename)
        self._generate_xlsx(excel_data, resultado_final, xlsx_filename)
        
        return {
            'pdf': pdf_filename,
            'xlsx': xlsx_filename
        }
        
    def _generate_pdf(self, table_data, resultado_final, filename):
        self.doc = SimpleDocTemplate(filename, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=5.0*cm, bottomMargin=4.5*cm)
        
        # Linha final com resultado
        table_data.append([
            Paragraph(f"<b>{self.t['final_result']}</b>", self.style_bold_left), '',
            Paragraph(f"<b>{format_currency(resultado_final)}</b>", self.style_bold_right)
        ])
            
        style = TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'), ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), self.style_normal.fontName), 
            ('FONTSIZE', (0, 0), (-1, -1), 7.6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0.856),
            ('TOPPADDING', (0, 0), (-1, -1), 0.856),
            ('LINEABOVE', (0, 0), (-1, 0), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, -1), (-1, -1), 1, colors.black, 1, None, None, 2),
        ])

        # Aplicar zebra apenas nos grupos (linhas com texto em negrito)
        row_count = 0
        for i, row_content in enumerate(table_data):
            if i == 0: continue  # Skip header
            if len(row_content) >= 1 and isinstance(row_content[0], Paragraph):
                if '<b>' in row_content[0].text and not any(keyword in row_content[0].text.upper() for keyword in ['RECEITA LÍQUIDA', 'LUCRO BRUTO', 'RESULTADO OPERACIONAL', 'LUCRO ANTES', 'RESULTADO DO EXERCÍCIO']):
                    if row_count % 2 == 1:
                        style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F0F0F0'))
                    row_count += 1

        col_widths = [10.5 * cm, 3.5 * cm, 3.5 * cm]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(style)
        self.story.append(table)
        
        admin_block, contador_block = self._create_assinaturas()
        assinaturas_table = Table([[admin_block, contador_block]], colWidths=[(A4[0]/2 - 2*cm), (A4[0]/2 - 2*cm)])
        assinaturas_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))

        self.story.append(Spacer(1, 4 * cm))
        self.story.append(assinaturas_table)

        self.doc.build(self.story, onFirstPage=self._header_footer, onLaterPages=self._header_footer)
        print(f"PDF '{filename}' gerado com sucesso.")

def gerar_dre(codi_emp, data_inicial, data_final, ingles=False):
    """Função principal para gerar DRE - VERSÃO CORRIGIDA"""
    lang_text = "inglês" if ingles else "português"
    print(f"Gerando DRE em {lang_text} - Empresa: {codi_emp}, Período: {data_inicial} a {data_final}")
    
    # Gerar timestamp único para esta execução
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    print(f"Timestamp da execução: {timestamp}")
    
    # Guarda a solicitação original de inglês para interface
    interface_ingles = ingles
    consulta_ingles = ingles
    
    # Primeiro, verificar se as tabelas de idiomas existem se inglês foi solicitado
    if ingles:
        try:
            with pyodbc.connect(CONN_STR) as conn:
                with conn.cursor() as cursor:
                    # Verificar se existe CTCONTAS_IDIOMAS
                    try:
                        cursor.execute(f"SELECT COUNT(*) FROM BETHADBA.CTCONTAS_IDIOMAS WHERE CODI_EMP = {codi_emp} AND I_IDIOMAS = 1")
                        contas_count = cursor.fetchone()[0]
                    except:
                        contas_count = 0
                    
                    # Verificar se existe CTGRUPOSDRE_IDIOMAS  
                    try:
                        cursor.execute(f"SELECT COUNT(*) FROM BETHADBA.CTGRUPOSDRE_IDIOMAS WHERE CODI_EMP = {codi_emp} AND I_IDIOMAS = 1")
                        grupos_count = cursor.fetchone()[0]
                    except:
                        grupos_count = 0
                    
                    if contas_count == 0 and grupos_count == 0:
                        print("Aviso: Não foram encontradas traduções em inglês no banco. Usando textos em português com interface em inglês.")
                        consulta_ingles = False  # Desabilita JOINs mas mantém interface em inglês
                    else:
                        print(f"Encontradas traduções: {contas_count} contas e {grupos_count} grupos em inglês.")
        except Exception as e:
            print(f"Erro ao verificar traduções: {e}. Prosseguindo sem traduções do banco.")
            consulta_ingles = False
    
    queries = get_queries(codi_emp, data_inicial, data_final, consulta_ingles)
    db_data = fetch_data(CONN_STR, queries)
    
    if db_data:
        # Salvar JSON com timestamp
        json_filename = salvar_dados_json(db_data, codi_emp, data_inicial, data_final, interface_ingles, timestamp)
        
        # Criar generator com timestamp
        generator = DREGenerator(db_data, codi_emp, data_inicial, data_final, interface_ingles, timestamp)
        generated_files = generator.run()
        
        # Compilar todos os arquivos gerados
        all_files = {
            'json': json_filename,
            'pdf': generated_files['pdf'],
            'xlsx': generated_files['xlsx']
        }
        
        print(f"\n📁 Arquivos gerados com sucesso:")
        print(f"   • JSON: {all_files['json']}")
        print(f"   • PDF:  {all_files['pdf']}")
        print(f"   • XLSX: {all_files['xlsx']}")
        
        # Renomear arquivos PDF e XLSX após geração
        try:
            renamed_pdf = file_renamer.rename_file_after_generation(all_files['pdf'], codi_emp, data_inicial, data_final, timestamp, ingles, conn_str=CONN_STR)
            renamed_xlsx = file_renamer.rename_file_after_generation(all_files['xlsx'], codi_emp, data_inicial, data_final, timestamp, ingles, conn_str=CONN_STR)

            all_files['pdf'] = renamed_pdf
            all_files['xlsx'] = renamed_xlsx

            print(f"   • PDF renomeado:  {renamed_pdf}")
            print(f"   • XLSX renomeado: {renamed_xlsx}")
        except Exception as e:
            print(f"Aviso: Erro ao renomear arquivos: {e}. Mantendo nomes originais.")
        
        return all_files
    else:
        print("Erro ao obter dados do banco.")
        return None


# EXEMPLO DE USO
if __name__ == "__main__":
    gerar_dre(2286, "2025-05-01", "2025-05-31")