import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, timedelta
from collections import OrderedDict
from file_renamer import rename_file_after_generation

# --- 1. CONFIGURAÇÃO E FUNÇÕES AUXILIARES ---
CONN_STR = (
    "DRIVER={SQL Anywhere 17};"
    "HOST=NOTE-GO-273.go.local:2638;"
    "DBN=contabil;"
    "UID=ESTATISTICA002;"
    "PWD=U0T/wq6OdZ0oYSpvJRWGfg==;"
)

def gerar_periodos_mensais(data_inicio, data_fim):
    """
    Gera uma lista de períodos mensais entre duas datas
    """
    inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
    fim = datetime.strptime(data_fim, '%Y-%m-%d')
    
    periodos = []
    current = inicio.replace(day=1)
    
    while current <= fim:
        primeiro_dia = current
        
        if current.month == 12:
            next_month = current.replace(year=current.year + 1, month=1, day=1)
        else:
            next_month = current.replace(month=current.month + 1, day=1)
        
        ultimo_dia = next_month - timedelta(days=1)
        
        if ultimo_dia.date() > fim.date():
            ultimo_dia = fim
        
        if primeiro_dia.date() < inicio.date():
            primeiro_dia = inicio
            
        label_mes = current.strftime('%m/%Y')
        
        periodos.append((
            primeiro_dia.strftime('%Y-%m-%d'),
            ultimo_dia.strftime('%Y-%m-%d'),
            label_mes
        ))
        
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)
    
    return periodos

def format_date_br(date_str):
    """
    Converte data do formato ISO (YYYY-MM-DD) para formato brasileiro (DD-MM-YYYY)
    """
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        return date_obj.strftime('%d-%m-%Y')
    except ValueError:
        return date_str

def get_sql_queries(codigo_empresa, data_inicio, data_fim, ingles=False):
    """
    Gera as consultas SQL com parâmetros dinâmicos
    """
    if ingles:
        nome_cta_select = "COALESCE(ci.DESCRICAO, c.NOME_CTA) AS NOME_CTA"
        join_idioma = "LEFT JOIN BETHADBA.CTCONTAS_IDIOMAS ci ON ci.CODI_EMP = c.CODI_EMP AND ci.CODI_CTA = c.CODI_CTA AND ci.I_IDIOMAS = 1"
    else:
        nome_cta_select = "c.NOME_CTA"
        join_idioma = ""

    return {
        "info_empresa": f"SELECT NOME_EMP, CGCE_EMP FROM BETHADBA.GEEMPRE WHERE CODI_EMP = {codigo_empresa}",
        "dados_relatorio": f"""
            WITH Movimentos AS (
                SELECT 
                    CDEB_LAN AS CODI_CTA, 
                    DATA_LAN, 
                    VLOR_LAN AS VALOR_DEBITO,
                    0 AS VALOR_CREDITO
                FROM BETHADBA.CTLANCTO 
                WHERE CODI_EMP = {codigo_empresa}
                UNION ALL
                SELECT 
                    CCRE_LAN AS CODI_CTA, 
                    DATA_LAN, 
                    0 AS VALOR_DEBITO,
                    VLOR_LAN AS VALOR_CREDITO
                FROM BETHADBA.CTLANCTO 
                WHERE CODI_EMP = {codigo_empresa}
            ),
            SaldosAnteriores AS (
                SELECT 
                    CODI_CTA, 
                    0 AS SALDOANT
                FROM Movimentos
                GROUP BY CODI_CTA
            ),
            MovimentosPeriodo AS (
                SELECT
                    CODI_CTA,
                    SUM(VALOR_DEBITO) AS TOTDEB,
                    SUM(VALOR_CREDITO) AS TOTCRE
                FROM Movimentos
                WHERE DATA_LAN BETWEEN '{data_inicio}' AND '{data_fim}'
                GROUP BY CODI_CTA
            )
            SELECT 
                c.CODI_CTA, 
                {nome_cta_select}, 
                c.CLAS_CTA, 
                c.TIPO_CTA,
                0 AS SALDOANT,
                COALESCE(mp.TOTDEB, 0) AS TOTDEB,
                COALESCE(mp.TOTCRE, 0) AS TOTCRE
            FROM BETHADBA.CTCONTAS c
            {join_idioma}
            LEFT JOIN MovimentosPeriodo mp ON c.CODI_CTA = mp.CODI_CTA
            WHERE c.CODI_EMP = {codigo_empresa} 
            AND (c.CLAS_CTA LIKE '3%' OR c.CLAS_CTA LIKE '4%' OR c.CLAS_CTA LIKE '5%')
            ORDER BY c.CLAS_CTA
        """
    }

def get_sql_queries_mensal(codigo_empresa, data_inicio_mes, data_fim_mes, ingles=False):
    """
    Gera consulta SQL para um mês específico
    """
    if ingles:
        nome_cta_select = "COALESCE(ci.DESCRICAO, c.NOME_CTA) AS NOME_CTA"
        join_idioma = "LEFT JOIN BETHADBA.CTCONTAS_IDIOMAS ci ON ci.CODI_EMP = c.CODI_EMP AND ci.CODI_CTA = c.CODI_CTA AND ci.I_IDIOMAS = 1"
    else:
        nome_cta_select = "c.NOME_CTA"
        join_idioma = ""

    return f"""
        WITH Movimentos AS (
            SELECT 
                CDEB_LAN AS CODI_CTA, 
                DATA_LAN, 
                VLOR_LAN AS VALOR_DEBITO,
                0 AS VALOR_CREDITO
            FROM BETHADBA.CTLANCTO 
            WHERE CODI_EMP = {codigo_empresa}
            UNION ALL
            SELECT 
                CCRE_LAN AS CODI_CTA, 
                DATA_LAN, 
                0 AS VALOR_DEBITO,
                VLOR_LAN AS VALOR_CREDITO
            FROM BETHADBA.CTLANCTO 
            WHERE CODI_EMP = {codigo_empresa}
        ),
        MovimentosMes AS (
            SELECT
                CODI_CTA,
                SUM(VALOR_DEBITO) AS TOTDEB,
                SUM(VALOR_CREDITO) AS TOTCRE
            FROM Movimentos
            WHERE DATA_LAN BETWEEN '{data_inicio_mes}' AND '{data_fim_mes}'
            GROUP BY CODI_CTA
        )
        SELECT 
            c.CODI_CTA, 
            {nome_cta_select}, 
            c.CLAS_CTA, 
            c.TIPO_CTA,
            COALESCE(mm.TOTDEB, 0) AS TOTDEB,
            COALESCE(mm.TOTCRE, 0) AS TOTCRE
        FROM BETHADBA.CTCONTAS c
        {join_idioma}
        LEFT JOIN MovimentosMes mm ON c.CODI_CTA = mm.CODI_CTA
        WHERE c.CODI_EMP = {codigo_empresa}
        AND (c.CLAS_CTA LIKE '3%' OR c.CLAS_CTA LIKE '4%' OR c.CLAS_CTA LIKE '5%')
        ORDER BY c.CLAS_CTA
    """

# --- 2. ACESSO AO BANCO DE DADOS ---
def execute_query(query_key, sql_queries):
    """
    Executa uma consulta SQL específica
    """
    try:
        with pyodbc.connect(CONN_STR) as conn:
            with conn.cursor() as cursor:
                print(f"Executando query '{query_key}'...")
                cursor.execute(sql_queries[query_key])
                rows = cursor.fetchall()
                print(f"Query '{query_key}' retornou {len(rows)} linhas.")
                return rows
    except pyodbc.Error as ex:
        sqlstate = ex.args[0]
        print(f"Erro de Banco de Dados ao executar query '{query_key}': {sqlstate}")
        print(ex)
        return None

# --- 3. PROCESSAMENTO DE DADOS (MULTI-MESES) ---
def process_financial_data_multimonths(codigo_empresa, data_inicio, data_fim, ingles=False):
    """
    Processa dados financeiros para múltiplos meses
    """
    periodos_mensais = gerar_periodos_mensais(data_inicio, data_fim)
    
    saldos_anteriores = {}
    
    dados_por_mes = {}
    accounts = OrderedDict()
    
    for inicio_mes, fim_mes, label_mes in periodos_mensais:
        print(f"Processando período: {label_mes}")
        
        query_mes = get_sql_queries_mensal(codigo_empresa, inicio_mes, fim_mes, ingles)
        try:
            with pyodbc.connect(CONN_STR) as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query_mes)
                    rows = cursor.fetchall()
                    dados_por_mes[label_mes] = rows
                    
                    for row in rows:
                        codigo = row[0]
                        descricao = row[1].strip()
                        class_code = row[2].strip()
                        tipo_conta = row[3].strip()
                        total_debito = Decimal(str(row[4]))
                        total_credito = Decimal(str(row[5]))
                        
                        if not class_code.startswith(('3', '4', '5')):
                            continue
                        
                        if class_code not in accounts:
                            accounts[class_code] = {
                                'codigo': codigo,
                                'classificacao': class_code,
                                'descricao': descricao,
                                'is_synthetic': tipo_conta == 'S',
                                'saldo_anterior': Decimal(0),
                                'movimentos_mensais': {},
                                'saldo_acumulado': Decimal(0)
                            }
                        
                        valor_mes = total_debito - total_credito
                        
                        if class_code.startswith('3'):
                            valor_mes = -valor_mes
                        
                        accounts[class_code]['movimentos_mensais'][label_mes] = valor_mes
                        
        except pyodbc.Error as ex:
            print(f"Erro ao buscar dados do mês {label_mes}: {ex}")
            dados_por_mes[label_mes] = []
    
    for code in sorted(accounts.keys()):
        if accounts[code]['is_synthetic']:
            accounts[code]['saldo_anterior'] = Decimal(0)
            accounts[code]['movimentos_mensais'] = {label: Decimal(0) for _, _, label in periodos_mensais}
    
    synthetic_accounts = [code for code in accounts.keys() if accounts[code]['is_synthetic']]
    synthetic_accounts.sort(key=len, reverse=True)
    
    for synthetic_code in synthetic_accounts:
        total_saldo_anterior = Decimal(0)
        totais_mensais = {label: Decimal(0) for _, _, label in periodos_mensais}
        
        for account_code in accounts.keys():
            if (account_code.startswith(synthetic_code) and 
                account_code != synthetic_code and 
                not accounts[account_code]['is_synthetic']):
                
                total_saldo_anterior += accounts[account_code]['saldo_anterior']
                
                for label_mes in totais_mensais:
                    if label_mes in accounts[account_code]['movimentos_mensais']:
                        totais_mensais[label_mes] += accounts[account_code]['movimentos_mensais'][label_mes]
        
        accounts[synthetic_code]['saldo_anterior'] = total_saldo_anterior
        accounts[synthetic_code]['movimentos_mensais'] = totais_mensais
    
    for code in accounts:
        saldo_atual = accounts[code]['saldo_anterior']
        for _, _, label_mes in periodos_mensais:
            if label_mes in accounts[code]['movimentos_mensais']:
                saldo_atual += accounts[code]['movimentos_mensais'][label_mes]
        accounts[code]['saldo_acumulado'] = saldo_atual
    
    contas_antes = len(accounts)
    accounts_filtered = OrderedDict()
    for code, account in accounts.items():
        tem_movimento = False
        
        for _, _, label_mes in periodos_mensais:
            if label_mes in account['movimentos_mensais']:
                if account['movimentos_mensais'][label_mes] != Decimal(0):
                    tem_movimento = True
                    break
        
        if account['saldo_acumulado'] != Decimal(0):
            tem_movimento = True
        
        if tem_movimento:
            accounts_filtered[code] = account
    
    accounts = accounts_filtered
    contas_depois = len(accounts)
    print(f"Filtro aplicado: {contas_antes - contas_depois} contas zeradas removidas. Restam {contas_depois} contas.")
    
    report_list = []
    for code in sorted(accounts.keys()):
        if code in accounts:
            report_list.append(accounts[code])
    
    resumo = {
        "RECEITAS": Decimal(0),
        "CUSTOS DAS VENDAS": Decimal(0), 
        "DESPESAS OPERACIONAIS": Decimal(0),
    }
    
    if "3" in accounts:
        for _, _, label_mes in periodos_mensais:
            if label_mes in accounts["3"]["movimentos_mensais"]:
                resumo["RECEITAS"] += accounts["3"]["movimentos_mensais"][label_mes]
    
    if "4" in accounts:
        for _, _, label_mes in periodos_mensais:
            if label_mes in accounts["4"]["movimentos_mensais"]:
                resumo["CUSTOS DAS VENDAS"] += accounts["4"]["movimentos_mensais"][label_mes]
    
    if "5" in accounts:
        for _, _, label_mes in periodos_mensais:
            if label_mes in accounts["5"]["movimentos_mensais"]:
                resumo["DESPESAS OPERACIONAIS"] += accounts["5"]["movimentos_mensais"][label_mes]
    
    return report_list, resumo, periodos_mensais, dados_por_mes

# --- 4. GERAÇÃO DO EXCEL (MULTI-MESES) ---
class ReportExcelMultiMonths:
    def __init__(self, filename, company_info, data, resumo, data_inicio, data_fim, periodos_mensais, ingles=False):
        self.filename = filename
        self.company_info = company_info
        self.data = data
        self.resumo = resumo
        self.data_inicio = data_inicio
        self.data_fim = data_fim
        self.periodos_mensais = periodos_mensais
        self.ingles = ingles
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Comparativo de Movimento"
        
        self.ws.sheet_view.showGridLines = False
        
        self.setup_styles()
        
        self.current_row = 1

    def setup_styles(self):
        self.font_header = Font(name='Arial', size=8, bold=True)
        self.font_title = Font(name='Arial', size=10, bold=True)
        self.font_normal = Font(name='Arial', size=8)
        self.font_synthetic = Font(name='Arial', size=8, bold=True)
        self.font_small = Font(name='Arial', size=7)
        self.fill_gray = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        self.align_left = Alignment(horizontal='left', vertical='center')
        self.align_right = Alignment(horizontal='right', vertical='center')
        self.align_center = Alignment(horizontal='center', vertical='center')
        thin_border = Side(border_style='thin', color='000000')
        self.border_all = Border(top=thin_border, bottom=thin_border, left=thin_border, right=thin_border)
        self.border_top_bottom = Border(top=thin_border, bottom=thin_border)

    def format_classification(self, classification):
        c = classification
        if len(c) >= 12: return f"{c[0]}.{c[1]}.{c[2]}.{c[3:6]}.{c[6:9]}.{c[9:]}"
        if len(c) >= 9: return f"{c[0]}.{c[1]}.{c[2]}.{c[3:6]}.{c[6:]}"
        if len(c) >= 6: return f"{c[0]}.{c[1]}.{c[2]}.{c[3:]}"
        if len(c) >= 3: return f"{c[0]}.{c[1]}.{c[2]}"
        if len(c) >= 2: return f"{c[0]}.{c[1]}"
        return c

    def format_currency(self, value, class_code=''):
        value = Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if value == 0: return "0,00"
        char = ''
        show_value = abs(value)
        if class_code.startswith('3'): char = 'C' if value > 0 else 'D'
        elif class_code.startswith(('4', '5')): char = 'D' if value > 0 else 'C'
        elif class_code == 'RESUMO_D': char = 'D'
        elif class_code == 'RESUMO_C': char = 'C'
        formatted_value = f"{show_value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"{formatted_value}{char}"

    def write_header(self):
        # ALTERAÇÃO: Tradução dos textos do cabeçalho
        empresa_lbl = "Company" if self.ingles else "Empresa"
        periodo_lbl = "Period" if self.ingles else "Período"
        folha_lbl = "Page" if self.ingles else "Folha"
        emissao_lbl = "Date" if self.ingles else "Emissão"
        hora_lbl = "Time" if self.ingles else "Hora"
        titulo_relatorio = "MOVEMENT COMPARISON" if self.ingles else "COMPARATIVO DE MOVIMENTO"

        self.ws[f'A{self.current_row}'] = f"{empresa_lbl}:"
        self.ws[f'A{self.current_row}'].font = self.font_normal
        self.ws[f'B{self.current_row}'] = self.company_info.get('nome', 'N/A')
        
        ultima_coluna_letra = get_column_letter(5 + len(self.periodos_mensais))
        
        self.ws[f'{chr(ord(ultima_coluna_letra)-1)}{self.current_row}'] = f"{folha_lbl}:"
        self.ws[f'{chr(ord(ultima_coluna_letra)-1)}{self.current_row}'].font = self.font_normal
        self.ws[f'{chr(ord(ultima_coluna_letra)-1)}{self.current_row}'].alignment = self.align_right
        self.ws[f'{ultima_coluna_letra}{self.current_row}'] = "0001"
        self.ws[f'{ultima_coluna_letra}{self.current_row}'].font = self.font_normal
        self.ws[f'{ultima_coluna_letra}{self.current_row}'].alignment = self.align_right
        
        self.current_row += 1
        
        self.ws[f'A{self.current_row}'] = "C.N.P.J.:"
        self.ws[f'B{self.current_row}'] = self.company_info.get('cnpj', 'N/A')
        
        self.ws[f'{chr(ord(ultima_coluna_letra)-1)}{self.current_row}'] = f"{emissao_lbl}:"
        self.ws[f'{chr(ord(ultima_coluna_letra)-1)}{self.current_row}'].font = self.font_normal
        self.ws[f'{chr(ord(ultima_coluna_letra)-1)}{self.current_row}'].alignment = self.align_right
        self.ws[f'{ultima_coluna_letra}{self.current_row}'] = "13/08/2025"
        self.ws[f'{ultima_coluna_letra}{self.current_row}'].font = self.font_normal
        self.ws[f'{ultima_coluna_letra}{self.current_row}'].alignment = self.align_right
        
        self.current_row += 1
        
        self.ws[f'A{self.current_row}'] = f"{periodo_lbl}:"
        self.ws[f'B{self.current_row}'] = f"{format_date_br(self.data_inicio)} - {format_date_br(self.data_fim)}"
        
        self.ws[f'{chr(ord(ultima_coluna_letra)-1)}{self.current_row}'] = f"{hora_lbl}:"
        self.ws[f'{chr(ord(ultima_coluna_letra)-1)}{self.current_row}'].font = self.font_normal
        self.ws[f'{chr(ord(ultima_coluna_letra)-1)}{self.current_row}'].alignment = self.align_right
        self.ws[f'{ultima_coluna_letra}{self.current_row}'] = "09:11:11"
        self.ws[f'{ultima_coluna_letra}{self.current_row}'].font = self.font_normal
        self.ws[f'{ultima_coluna_letra}{self.current_row}'].alignment = self.align_right
        
        self.current_row += 2
        
        self.ws[f'C{self.current_row}'] = titulo_relatorio
        self.ws[f'C{self.current_row}'].font = self.font_title
        self.ws[f'C{self.current_row}'].alignment = self.align_center
        
        self.current_row += 2
        
        if self.ingles:
            headers = ["Code", "Classification", "Description"]
            headers.extend([label for _, _, label in self.periodos_mensais])
            headers.append("Accumulated Balance")
        else:
            headers = ["Código", "Classificação", "Descrição"]
            headers.extend([label for _, _, label in self.periodos_mensais])
            headers.append("Saldo acumulado")
        
        for i, header in enumerate(headers):
            col_letter = get_column_letter(i + 1)
            cell = self.ws[f'{col_letter}{self.current_row}']
            cell.value = header
            cell.font = self.font_header
            cell.border = self.border_top_bottom
            if i >= 3:
                cell.alignment = self.align_right
            else:
                cell.alignment = self.align_left
        
        self.current_row += 1

    def write_resumo(self):
        self.current_row += 2
        
        # ALTERAÇÃO: Tradução do resumo
        resumo_titulo = "SUMMARY OF THE TRIAL BALANCE SHEET" if self.ingles else "RESUMO DO BALANCETE"
        self.ws[f'A{self.current_row}'] = resumo_titulo
        self.ws[f'A{self.current_row}'].font = Font(name='Arial', size=9, bold=True)
        
        num_cols = 3 + len(self.periodos_mensais) + 1
        for i in range(num_cols):
            self.ws[f'{get_column_letter(i + 1)}{self.current_row}'].border = Border(bottom=Side(border_style='thin'))
        
        self.current_row += 1
        
        receitas_mensais = {}
        custos_mensais = {}
        despesas_mensais = {}
        
        for _, _, label_mes in self.periodos_mensais:
            receitas_mensais[label_mes] = Decimal(0)
            custos_mensais[label_mes] = Decimal(0) 
            despesas_mensais[label_mes] = Decimal(0)
        
        for item in self.data:
            if item['classificacao'] == '3':
                for _, _, label_mes in self.periodos_mensais:
                    receitas_mensais[label_mes] = item['movimentos_mensais'].get(label_mes, Decimal(0))
            elif item['classificacao'] == '4':
                for _, _, label_mes in self.periodos_mensais:
                    custos_mensais[label_mes] = item['movimentos_mensais'].get(label_mes, Decimal(0))
            elif item['classificacao'] == '5':
                for _, _, label_mes in self.periodos_mensais:
                    despesas_mensais[label_mes] = item['movimentos_mensais'].get(label_mes, Decimal(0))
        
        # ALTERAÇÃO: Tradução dos itens do resumo
        if self.ingles:
            items = [
                ("ASSETS", {label: Decimal(0) for _, _, label in self.periodos_mensais}, Decimal(0), 'RESUMO_D'), 
                ("LIABILITIES", {label: Decimal(0) for _, _, label in self.periodos_mensais}, Decimal(0), 'RESUMO_D'),
                ("REVENUES", receitas_mensais, abs(self.resumo["RECEITAS"]), 'RESUMO_C'),
                ("COST OF SALES", custos_mensais, self.resumo["CUSTOS DAS VENDAS"], 'RESUMO_D'),
                ("OPERATING EXPENSES", despesas_mensais, self.resumo["DESPESAS OPERACIONAIS"], 'RESUMO_D'),
                ("NET REVENUES - PRODUCTS AND SERVICES - SPC", {label: Decimal(0) for _, _, label in self.periodos_mensais}, Decimal(0), 'RESUMO_D'),
                ("PAYROLL COSTS - SPC", {label: Decimal(0) for _, _, label in self.periodos_mensais}, Decimal(0), 'RESUMO_D'),
                ("CALCULATION RESULT - TRANSITORY", {label: Decimal(0) for _, _, label in self.periodos_mensais}, Decimal(0), 'RESUMO_D'),
                ("CALCULATION RESULT - TRANSITORY", {label: Decimal(0) for _, _, label in self.periodos_mensais}, Decimal(0), 'RESUMO_D'),
            ]
        else:
            items = [
                ("ATIVO", {label: Decimal(0) for _, _, label in self.periodos_mensais}, Decimal(0), 'RESUMO_D'), 
                ("PASSIVO", {label: Decimal(0) for _, _, label in self.periodos_mensais}, Decimal(0), 'RESUMO_D'),
                ("RECEITAS", receitas_mensais, abs(self.resumo["RECEITAS"]), 'RESUMO_C'),
                ("CUSTOS DAS VENDAS", custos_mensais, self.resumo["CUSTOS DAS VENDAS"], 'RESUMO_D'),
                ("DESPESAS OPERACIONAIS", despesas_mensais, self.resumo["DESPESAS OPERACIONAIS"], 'RESUMO_D'),
                ("RECEITAS LIQUIDAS DOS PRODUTOS E SERVICOS - SCP", {label: Decimal(0) for _, _, label in self.periodos_mensais}, Decimal(0), 'RESUMO_D'),
                ("DESPESAS OPERACIONAIS -SCP", {label: Decimal(0) for _, _, label in self.periodos_mensais}, Decimal(0), 'RESUMO_D'),
                ("APURACAO DE RESULTADO - TRANSITORIA", {label: Decimal(0) for _, _, label in self.periodos_mensais}, Decimal(0), 'RESUMO_D'),
                ("APURACAO DE RESULTADO - TRANSITORIA", {label: Decimal(0) for _, _, label in self.periodos_mensais}, Decimal(0), 'RESUMO_D'),
            ]
        
        ultima_coluna_letra = get_column_letter(3 + len(self.periodos_mensais) + 1)
        
        for idx, (desc, valores_mensais, valor_total, code) in enumerate(items):
            if idx % 2 != 0:
                for i in range(num_cols):
                    self.ws[f'{get_column_letter(i + 1)}{self.current_row}'].fill = self.fill_gray
            
            self.ws[f'A{self.current_row}'] = desc
            self.ws[f'A{self.current_row}'].font = self.font_normal
            
            col_index = 4
            for _, _, label_mes in self.periodos_mensais:
                col_letter = get_column_letter(col_index)
                valor_mes = valores_mensais.get(label_mes, Decimal(0))
                self.ws[f'{col_letter}{self.current_row}'] = self.format_currency(abs(valor_mes), code)
                self.ws[f'{col_letter}{self.current_row}'].font = self.font_normal
                self.ws[f'{col_letter}{self.current_row}'].alignment = self.align_right
                col_index += 1
            
            self.ws[f'{ultima_coluna_letra}{self.current_row}'] = self.format_currency(valor_total, code)
            self.ws[f'{ultima_coluna_letra}{self.current_row}'].font = self.font_normal
            self.ws[f'{ultima_coluna_letra}{self.current_row}'].alignment = self.align_right
            
            self.current_row += 1
        
        total_devedor = self.resumo["CUSTOS DAS VENDAS"] + self.resumo["DESPESAS OPERACIONAIS"]
        total_credor = abs(self.resumo["RECEITAS"])
        resultado_mes = total_credor - total_devedor

        self.current_row += 1
        
        # ALTERAÇÃO: Tradução dos totais
        contas_devedoras_lbl = "Debtor Accounts" if self.ingles else "CONTAS DEVEDORAS"
        contas_credoras_lbl = "Creditor Accounts" if self.ingles else "CONTAS CREDORAS"
        resultado_mes_lbl = "Month Profit/Loss" if self.ingles else "RESULTADO DO MES"
        resultado_exercicio_lbl = "Period Profit/Loss" if self.ingles else "RESULTADO DO EXERCÍCIO"
        
        if self.current_row % 2 == 0:
            for i in range(num_cols):
                self.ws[f'{get_column_letter(i + 1)}{self.current_row}'].fill = self.fill_gray
        
        self.ws[f'A{self.current_row}'] = contas_devedoras_lbl
        self.ws[f'A{self.current_row}'].font = self.font_normal
        
        col_index = 4
        for _, _, label_mes in self.periodos_mensais:
            col_letter = get_column_letter(col_index)
            total_devedor_mes = custos_mensais[label_mes] + despesas_mensais[label_mes]
            self.ws[f'{col_letter}{self.current_row}'] = self.format_currency(total_devedor_mes, 'RESUMO_D')
            self.ws[f'{col_letter}{self.current_row}'].font = self.font_normal
            self.ws[f'{col_letter}{self.current_row}'].alignment = self.align_right
            col_index += 1
        
        self.ws[f'{ultima_coluna_letra}{self.current_row}'] = self.format_currency(total_devedor, 'RESUMO_D')
        self.ws[f'{ultima_coluna_letra}{self.current_row}'].font = self.font_normal
        self.ws[f'{ultima_coluna_letra}{self.current_row}'].alignment = self.align_right
        
        self.current_row += 1
        
        if self.current_row % 2 == 0:
            for i in range(num_cols):
                self.ws[f'{get_column_letter(i + 1)}{self.current_row}'].fill = self.fill_gray
        
        self.ws[f'A{self.current_row}'] = contas_credoras_lbl
        self.ws[f'A{self.current_row}'].font = self.font_normal
        
        col_index = 4
        for _, _, label_mes in self.periodos_mensais:
            col_letter = get_column_letter(col_index)
            total_credor_mes = abs(receitas_mensais[label_mes])
            self.ws[f'{col_letter}{self.current_row}'] = self.format_currency(total_credor_mes, 'RESUMO_C')
            self.ws[f'{col_letter}{self.current_row}'].font = self.font_normal
            self.ws[f'{col_letter}{self.current_row}'].alignment = self.align_right
            col_index += 1
        
        self.ws[f'{ultima_coluna_letra}{self.current_row}'] = self.format_currency(total_credor, 'RESUMO_C')
        self.ws[f'{ultima_coluna_letra}{self.current_row}'].font = self.font_normal
        self.ws[f'{ultima_coluna_letra}{self.current_row}'].alignment = self.align_right

        self.current_row += 2
        
        if self.current_row % 2 == 0:
            for i in range(num_cols):
                self.ws[f'{get_column_letter(i + 1)}{self.current_row}'].fill = self.fill_gray
        
        self.ws[f'A{self.current_row}'] = resultado_mes_lbl
        self.ws[f'A{self.current_row}'].font = self.font_normal
        
        col_index = 4
        for _, _, label_mes in self.periodos_mensais:
            col_letter = get_column_letter(col_index)
            resultado_mes_individual = abs(receitas_mensais[label_mes]) - custos_mensais[label_mes] - despesas_mensais[label_mes]
            self.ws[f'{col_letter}{self.current_row}'] = self.format_currency(resultado_mes_individual, 'RESUMO_C' if resultado_mes_individual > 0 else 'RESUMO_D')
            self.ws[f'{col_letter}{self.current_row}'].font = self.font_normal
            self.ws[f'{col_letter}{self.current_row}'].alignment = self.align_right
            col_index += 1
        
        self.ws[f'{ultima_coluna_letra}{self.current_row}'] = self.format_currency(resultado_mes, 'RESUMO_C' if resultado_mes > 0 else 'RESUMO_D')
        self.ws[f'{ultima_coluna_letra}{self.current_row}'].font = self.font_normal
        self.ws[f'{ultima_coluna_letra}{self.current_row}'].alignment = self.align_right
        
        self.current_row += 1
        
        if self.current_row % 2 == 0:
            for i in range(num_cols):
                self.ws[f'{get_column_letter(i + 1)}{self.current_row}'].fill = self.fill_gray
        
        self.ws[f'A{self.current_row}'] = resultado_exercicio_lbl
        self.ws[f'A{self.current_row}'].font = self.font_normal
        
        col_index = 4
        for _, _, label_mes in self.periodos_mensais:
            col_letter = get_column_letter(col_index)
            resultado_mes_individual = abs(receitas_mensais[label_mes]) - custos_mensais[label_mes] - despesas_mensais[label_mes]
            self.ws[f'{col_letter}{self.current_row}'] = self.format_currency(resultado_mes_individual, 'RESUMO_C' if resultado_mes_individual > 0 else 'RESUMO_D')
            self.ws[f'{col_letter}{self.current_row}'].font = self.font_normal
            self.ws[f'{col_letter}{self.current_row}'].alignment = self.align_right
            col_index += 1
        
        self.ws[f'{ultima_coluna_letra}{self.current_row}'] = self.format_currency(resultado_mes, 'RESUMO_C' if resultado_mes > 0 else 'RESUMO_D')
        self.ws[f'{ultima_coluna_letra}{self.current_row}'].font = self.font_normal
        self.ws[f'{ultima_coluna_letra}{self.current_row}'].alignment = self.align_right
        
        self.current_row += 3
        
        self.ws[f'A{self.current_row}'] = "Sistema licenciado para GF SERVICOS DE CONTABILIDADE S/S"
        self.ws[f'A{self.current_row}'].font = self.font_small

    def write_data_row(self, item, use_alternating_color=False):
        if use_alternating_color:
            num_cols = 3 + len(self.periodos_mensais) + 1
            for i in range(num_cols):
                self.ws[f'{get_column_letter(i + 1)}{self.current_row}'].fill = self.fill_gray
        
        font = self.font_synthetic if item['is_synthetic'] else self.font_normal
        
        self.ws[f'A{self.current_row}'].value = str(item['codigo'])
        self.ws[f'A{self.current_row}'].font = font
        self.ws[f'A{self.current_row}'].alignment = self.align_left
        
        self.ws[f'B{self.current_row}'].value = self.format_classification(item['classificacao'])
        self.ws[f'B{self.current_row}'].font = font
        self.ws[f'B{self.current_row}'].alignment = self.align_left
        
        self.ws[f'C{self.current_row}'].value = item['descricao']
        self.ws[f'C{self.current_row}'].font = font
        self.ws[f'C{self.current_row}'].alignment = self.align_left
        
        col_index = 4
        for _, _, label_mes in self.periodos_mensais:
            col_letter = get_column_letter(col_index)
            cell = self.ws[f'{col_letter}{self.current_row}']
            valor_mes = item['movimentos_mensais'].get(label_mes, Decimal(0))
            cell.value = self.format_currency(valor_mes, item['classificacao'])
            cell.font = font
            cell.alignment = self.align_right
            col_index += 1
        
        col_letter = get_column_letter(col_index)
        cell = self.ws[f'{col_letter}{self.current_row}']
        cell.value = self.format_currency(item['saldo_acumulado'], item['classificacao'])
        cell.font = font
        cell.alignment = self.align_right
        
        self.current_row += 1

    def adjust_columns(self):
        num_meses = len(self.periodos_mensais)
        column_widths = { 1: 8, 2: 20, 3: 45 }
        for i in range(4, 4 + num_meses):
            column_widths[i] = 15
        column_widths[4 + num_meses] = 15
        for col_num, width in column_widths.items():
            self.ws.column_dimensions[get_column_letter(col_num)].width = width
            
    def generate(self):
        self.write_header()
        # Congela o cabeçalho
        self.ws.freeze_panes = f'A{self.current_row}'
        for idx, item in enumerate(self.data):
            self.write_data_row(item, use_alternating_color=(idx % 2 != 0))
        self.write_resumo()
        self.adjust_columns()
        self.wb.save(self.filename)
        print(f"Relatório Excel multi-meses '{self.filename}' gerado com sucesso.")

def gerar_relatorio(codigo_empresa=124, data_inicio='2025-02-01', data_fim='2025-02-28', ingles=False):
    """
    Gera o relatório contábil com parâmetros customizáveis e retorna os nomes dos arquivos gerados.
    """
    # ALTERAÇÃO 1: Inicializa uma lista para armazenar os nomes dos arquivos e gera um timestamp
    arquivos_gerados = []
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    print(f"\n=== GERANDO RELATÓRIO ===")
    print(f"Empresa: {codigo_empresa}")
    print(f"Período: {data_inicio} a {data_fim}")
    print(f"Idioma: {'Inglês' if ingles else 'Português'}")
    print("="*50)
    

    # Sempre gerar relatório acumulado do ano: data_inicio = 1º de janeiro do ano da data_fim
    ano_competencia = data_fim[:4]
    data_inicio_acumulado = f"{ano_competencia}-01-01"
    print(f"Modo: Acumulado (multi-meses)")
    print(f"Período: {data_inicio_acumulado} a {data_fim}")

    periodos_mensais = gerar_periodos_mensais(data_inicio_acumulado, data_fim)

    sql_queries = get_sql_queries(codigo_empresa, data_inicio_acumulado, data_fim, ingles)

    info_result = execute_query("info_empresa", sql_queries)
    if info_result:
        nome_emp, cnpj_raw = info_result[0]
        company_info = {
            'nome': nome_emp.strip(),
            'cnpj': f"{cnpj_raw[:2]}.{cnpj_raw[2:5]}.{cnpj_raw[5:8]}/{cnpj_raw[8:12]}-{cnpj_raw[12:]}"
        }
    else:
        company_info = {'nome': 'EMPRESA NÃO ENCONTRADA', 'cnpj': '00.000.000/0000-00'}

    # Define language suffix for filename
    lang_suffix = '_EN' if ingles else ''

    print("\nProcessando dados para acumulado...")
    report_data_list, resumo_data, periodos_mensais, dados_por_mes = process_financial_data_multimonths(
        codigo_empresa, data_inicio_acumulado, data_fim, ingles
    )

    if not report_data_list:
        print("\n!!! ATENÇÃO: A lista de dados para o relatório está vazia após o processamento.")
    else:
        print(f"Gerando Excel acumulado com {len(report_data_list)} linhas...")

        filename = f"Comparativo_Movimento_{codigo_empresa}_{data_inicio.replace('-','')}_{data_fim.replace('-','')}{lang_suffix}_{timestamp}.xlsx"

        excel_generator = ReportExcelMultiMonths(
            filename, company_info, report_data_list, resumo_data, 
            data_inicio, data_fim, periodos_mensais, ingles
        )
        excel_generator.generate()
        arquivos_gerados.append(filename)

    # Renomear arquivos conforme padrão solicitado
    if arquivos_gerados:
        print(f"\n📝 Renomeando arquivos...")
        arquivos_renomeados = []
        for filename in arquivos_gerados:
            try:
                filename_renamed = rename_file_after_generation(
                    filename, str(codigo_empresa), data_inicio, data_fim,
                    timestamp, ingles, 'comparativo', CONN_STR
                )
                arquivos_renomeados.append(filename_renamed)
            except Exception as e:
                print(f"⚠️  Aviso: Erro ao renomear arquivo {filename}: {e}")
                arquivos_renomeados.append(filename)
        arquivos_gerados = arquivos_renomeados

    return arquivos_gerados

def process_financial_data_single_month(db_data):
    """
    Processa dados financeiros para um único mês (função original)
    """
    accounts = OrderedDict()
    
    for row in db_data:
        codigo = row[0]
        descricao = row[1].strip()
        class_code = row[2].strip()
        tipo_conta = row[3].strip()
        saldo_anterior = Decimal(str(row[4]))
        total_debito = Decimal(str(row[5]))
        total_credito = Decimal(str(row[6]))
        
        if not class_code.startswith(('3', '4', '5')):
            continue
        
        valor_mes = total_debito - total_credito
        
        if class_code.startswith('3'):
            valor_mes = -valor_mes
        
        saldo_acumulado = saldo_anterior + (total_debito - total_credito)
        
        if class_code.startswith('3'):
            saldo_acumulado = saldo_anterior - (total_debito - total_credito)
        
        accounts[class_code] = {
            'codigo': codigo,
            'classificacao': class_code,
            'descricao': descricao,
            'is_synthetic': tipo_conta == 'S',
            'valor_mes': valor_mes,
            'saldo_acumulado': saldo_acumulado,
            'total_debito': total_debito,
            'total_credito': total_credito
        }

    for code in sorted(accounts.keys()):
        if accounts[code]['is_synthetic']:
            accounts[code]['valor_mes'] = Decimal(0)
            accounts[code]['saldo_acumulado'] = Decimal(0)
            accounts[code]['total_debito'] = Decimal(0)
            accounts[code]['total_credito'] = Decimal(0)

    synthetic_accounts = [code for code in accounts.keys() if accounts[code]['is_synthetic']]
    synthetic_accounts.sort(key=len, reverse=True)
    
    for synthetic_code in synthetic_accounts:
        total_valor_mes = Decimal(0)
        total_saldo_acumulado = Decimal(0)
        total_debito = Decimal(0)
        total_credito = Decimal(0)
        
        for account_code in accounts.keys():
            if (account_code.startswith(synthetic_code) and 
                account_code != synthetic_code and 
                not accounts[account_code]['is_synthetic']):
                
                total_valor_mes += accounts[account_code]['valor_mes']
                total_saldo_acumulado += accounts[account_code]['saldo_acumulado']
                total_debito += accounts[account_code]['total_debito']
                total_credito += accounts[account_code]['total_credito']
        
        accounts[synthetic_code]['valor_mes'] = total_valor_mes
        accounts[synthetic_code]['saldo_acumulado'] = total_saldo_acumulado
        accounts[synthetic_code]['total_debito'] = total_debito
        accounts[synthetic_code]['total_credito'] = total_credito

    contas_antes = len(accounts)
    accounts_filtered = OrderedDict()
    for code, account in accounts.items():
        if account['valor_mes'] != Decimal(0) or account['saldo_acumulado'] != Decimal(0):
            accounts_filtered[code] = account
    
    accounts = accounts_filtered
    contas_depois = len(accounts)
    print(f"Filtro aplicado: {contas_antes - contas_depois} contas zeradas removidas. Restam {contas_depois} contas.")

    report_list = []
    for code in sorted(accounts.keys()):
        if code in accounts:
            report_list.append(accounts[code])
    
    resumo = {
        "RECEITAS": accounts.get("3", {}).get("valor_mes", Decimal(0)),
        "CUSTOS DAS VENDAS": accounts.get("4", {}).get("valor_mes", Decimal(0)),
        "DESPESAS OPERACIONAIS": accounts.get("5", {}).get("valor_mes", Decimal(0)),
    }
    
    return report_list, resumo

# Classe para mês único (simplificada)
class ReportExcelSingleMonth:
    def __init__(self, filename, company_info, data, resumo, data_inicio, data_fim, ingles=False):
        self.filename = filename
        self.company_info = company_info
        self.data = data
        self.resumo = resumo
        self.data_inicio = data_inicio
        self.data_fim = data_fim
        self.ingles = ingles
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Comparativo de Movimento"
        self.ws.sheet_view.showGridLines = False
        self.setup_styles()
        self.current_row = 1

    def setup_styles(self):
        self.font_header = Font(name='Arial', size=8, bold=True)
        self.font_title = Font(name='Arial', size=10, bold=True)
        self.font_normal = Font(name='Arial', size=8)
        self.font_synthetic = Font(name='Arial', size=8, bold=True)
        self.fill_gray = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        self.align_left = Alignment(horizontal='left', vertical='center')
        self.align_right = Alignment(horizontal='right', vertical='center')
        self.align_center = Alignment(horizontal='center', vertical='center')
        thin_border = Side(border_style='thin', color='000000')
        self.border_top_bottom = Border(top=thin_border, bottom=thin_border)

    def format_classification(self, classification):
        c = classification
        if len(c) >= 12: return f"{c[0]}.{c[1]}.{c[2]}.{c[3:6]}.{c[6:9]}.{c[9:]}"
        if len(c) >= 9: return f"{c[0]}.{c[1]}.{c[2]}.{c[3:6]}.{c[6:]}"
        if len(c) >= 6: return f"{c[0]}.{c[1]}.{c[2]}.{c[3:]}"
        if len(c) >= 3: return f"{c[0]}.{c[1]}.{c[2]}"
        if len(c) >= 2: return f"{c[0]}.{c[1]}"
        return c

    def format_currency(self, value, class_code=''):
        value = Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if value == 0: return "0,00"
        char = ''
        show_value = abs(value)
        if class_code.startswith('3'): char = 'C' if value > 0 else 'D'
        elif class_code.startswith(('4', '5')): char = 'D' if value > 0 else 'C'
        formatted_value = f"{show_value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"{formatted_value}{char}"

    def write_header(self):
        # ALTERAÇÃO: Tradução dos textos do cabeçalho
        empresa_lbl = "Company" if self.ingles else "Empresa"
        periodo_lbl = "Period" if self.ingles else "Período"
        folha_lbl = "Page" if self.ingles else "Folha"
        emissao_lbl = "Date" if self.ingles else "Emissão"
        hora_lbl = "Time" if self.ingles else "Hora"
        titulo_relatorio = "MOVEMENT COMPARISON" if self.ingles else "COMPARATIVO DE MOVIMENTO"

        self.ws[f'A{self.current_row}'] = f"{empresa_lbl}:"
        self.ws[f'B{self.current_row}'] = self.company_info.get('nome', 'N/A')
        self.ws[f'E{self.current_row}'] = f"{folha_lbl}: 0001"
        self.current_row += 1
        
        self.ws[f'A{self.current_row}'] = "C.N.P.J.:"
        self.ws[f'B{self.current_row}'] = self.company_info.get('cnpj', 'N/A')
        self.ws[f'E{self.current_row}'] = f"{emissao_lbl}: 13/08/2025"
        self.current_row += 1
        
        self.ws[f'A{self.current_row}'] = f"{periodo_lbl}:"
        self.ws[f'B{self.current_row}'] = f"{format_date_br(self.data_inicio)} - {format_date_br(self.data_fim)}"
        self.ws[f'E{self.current_row}'] = f"{hora_lbl}: 09:11:11"
        self.current_row += 2
        
        self.ws[f'C{self.current_row}'] = titulo_relatorio
        self.ws[f'C{self.current_row}'].font = self.font_title
        self.current_row += 2
        
        mes_ano = f"{self.data_inicio[5:7]}/{self.data_inicio[0:4]}"
        if self.ingles:
            headers = ["Code", "Classification", "Description", mes_ano, "Accumulated Balance"]
        else:
            headers = ["Código", "Classificação", "Descrição", mes_ano, "Saldo acumulado"]

        for i, header in enumerate(headers):
            col = get_column_letter(i + 1)
            self.ws[f'{col}{self.current_row}'] = header
            self.ws[f'{col}{self.current_row}'].font = self.font_header
            self.ws[f'{col}{self.current_row}'].border = self.border_top_bottom
        self.current_row += 1

    def write_data_row(self, item, use_alternating_color=False):
        if use_alternating_color:
            for i in range(5):
                self.ws[f'{get_column_letter(i + 1)}{self.current_row}'].fill = self.fill_gray
        
        font = self.font_synthetic if item['is_synthetic'] else self.font_normal
        
        self.ws[f'A{self.current_row}'] = str(item['codigo'])
        self.ws[f'A{self.current_row}'].font = font
        
        self.ws[f'B{self.current_row}'] = self.format_classification(item['classificacao'])
        self.ws[f'B{self.current_row}'].font = font
        
        self.ws[f'C{self.current_row}'] = item['descricao']
        self.ws[f'C{self.current_row}'].font = font
        
        self.ws[f'D{self.current_row}'] = self.format_currency(item['valor_mes'], item['classificacao'])
        self.ws[f'D{self.current_row}'].font = font
        self.ws[f'D{self.current_row}'].alignment = self.align_right
        
        self.ws[f'E{self.current_row}'] = self.format_currency(item['saldo_acumulado'], item['classificacao'])
        self.ws[f'E{self.current_row}'].font = font
        self.ws[f'E{self.current_row}'].alignment = self.align_right
        
        self.current_row += 1

    def generate(self):
        self.write_header()
        # Congela o cabeçalho
        self.ws.freeze_panes = f'A{self.current_row}'
        for idx, item in enumerate(self.data):
            self.write_data_row(item, idx % 2 != 0)
        
        self.write_resumo_single()
        
        for i, width in enumerate([8, 20, 45, 15, 15], 1):
            self.ws.column_dimensions[get_column_letter(i)].width = width
        self.wb.save(self.filename)
        print(f"Relatório Excel mês único '{self.filename}' gerado com sucesso.")

    def write_resumo_single(self):
        self.current_row += 2
        
        resumo_titulo = "SUMMARY OF THE TRIAL BALANCE SHEET" if self.ingles else "RESUMO DO BALANCETE"
        self.ws[f'A{self.current_row}'] = resumo_titulo
        self.ws[f'A{self.current_row}'].font = Font(name='Arial', size=9, bold=True)
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            self.ws[f'{col}{self.current_row}'].border = Border(bottom=Side(border_style='thin'))
        
        self.current_row += 1
        
        if self.ingles:
            items = [
                ("ASSETS", Decimal(0), 'RESUMO_D'), 
                ("LIABILITIES", Decimal(0), 'RESUMO_D'),
                ("REVENUES", abs(self.resumo["RECEITAS"]), 'RESUMO_C'),
                ("COST OF SALES", self.resumo["CUSTOS DAS VENDAS"], 'RESUMO_D'),
                ("OPERATING EXPENSES", self.resumo["DESPESAS OPERACIONAIS"], 'RESUMO_D'),
                ("NET REVENUES - PRODUCTS AND SERVICES - SPC", Decimal(0), 'RESUMO_D'),
                ("PAYROLL COSTS - SPC", Decimal(0), 'RESUMO_D'),
                ("CALCULATION RESULT - TRANSITORY", Decimal(0), 'RESUMO_D'),
                ("CALCULATION RESULT - TRANSITORY", Decimal(0), 'RESUMO_D'),
            ]
        else:
            items = [
                ("ATIVO", Decimal(0), 'RESUMO_D'), 
                ("PASSIVO", Decimal(0), 'RESUMO_D'),
                ("RECEITAS", abs(self.resumo["RECEITAS"]), 'RESUMO_C'),
                ("CUSTOS DAS VENDAS", self.resumo["CUSTOS DAS VENDAS"], 'RESUMO_D'),
                ("DESPESAS OPERACIONAIS", self.resumo["DESPESAS OPERACIONAIS"], 'RESUMO_D'),
                ("RECEITAS LIQUIDAS DOS PRODUTOS E SERVICOS - SCP", Decimal(0), 'RESUMO_D'),
                ("DESPESAS OPERACIONAIS -SCP", Decimal(0), 'RESUMO_D'),
                ("APURACAO DE RESULTADO - TRANSITORIA", Decimal(0), 'RESUMO_D'),
                ("APURACAO DE RESULTADO - TRANSITORIA", Decimal(0), 'RESUMO_D'),
            ]
        
        for idx, (desc, value, code) in enumerate(items):
            if idx % 2 != 0:
                for col in ['A', 'B', 'C', 'D', 'E']:
                    self.ws[f'{col}{self.current_row}'].fill = self.fill_gray
            
            self.ws[f'A{self.current_row}'] = desc
            self.ws[f'A{self.current_row}'].font = self.font_normal
            self.ws[f'E{self.current_row}'] = self.format_currency(value, code)
            self.ws[f'E{self.current_row}'].font = self.font_normal
            self.ws[f'E{self.current_row}'].alignment = self.align_right
            
            self.current_row += 1
        
        total_devedor = self.resumo["CUSTOS DAS VENDAS"] + self.resumo["DESPESAS OPERACIONAIS"]
        total_credor = abs(self.resumo["RECEITAS"])
        resultado_mes = total_credor - total_devedor

        self.current_row += 1
        
        contas_devedoras_lbl = "Debtor Accounts" if self.ingles else "CONTAS DEVEDORAS"
        contas_credoras_lbl = "Creditor Accounts" if self.ingles else "CONTAS CREDORAS"
        resultado_mes_lbl = "Month Profit/Loss" if self.ingles else "RESULTADO DO MES"
        resultado_exercicio_lbl = "Period Profit/Loss" if self.ingles else "RESULTADO DO EXERCÍCIO"
        
        self.ws[f'A{self.current_row}'] = contas_devedoras_lbl
        self.ws[f'A{self.current_row}'].font = self.font_normal
        self.ws[f'E{self.current_row}'] = self.format_currency(total_devedor, 'RESUMO_D')
        self.ws[f'E{self.current_row}'].font = self.font_normal
        self.ws[f'E{self.current_row}'].alignment = self.align_right
        
        self.current_row += 1
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            self.ws[f'{col}{self.current_row}'].fill = self.fill_gray
        
        self.ws[f'A{self.current_row}'] = contas_credoras_lbl
        self.ws[f'A{self.current_row}'].font = self.font_normal
        self.ws[f'E{self.current_row}'] = self.format_currency(total_credor, 'RESUMO_C')
        self.ws[f'E{self.current_row}'].font = self.font_normal
        self.ws[f'E{self.current_row}'].alignment = self.align_right

        self.current_row += 2
        
        self.ws[f'A{self.current_row}'] = resultado_mes_lbl
        self.ws[f'A{self.current_row}'].font = self.font_normal
        self.ws[f'E{self.current_row}'] = self.format_currency(resultado_mes, 'RESUMO_C' if resultado_mes > 0 else 'RESUMO_D')
        self.ws[f'E{self.current_row}'].font = self.font_normal
        self.ws[f'E{self.current_row}'].alignment = self.align_right
        
        self.current_row += 1
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            self.ws[f'{col}{self.current_row}'].fill = self.fill_gray
        
        self.ws[f'A{self.current_row}'] = resultado_exercicio_lbl
        self.ws[f'A{self.current_row}'].font = self.font_normal
        self.ws[f'E{self.current_row}'] = self.format_currency(resultado_mes, 'RESUMO_C' if resultado_mes > 0 else 'RESUMO_D')
        self.ws[f'E{self.current_row}'].font = self.font_normal
        self.ws[f'E{self.current_row}'].alignment = self.align_right
        
        self.current_row += 3
        
        self.ws[f'A{self.current_row}'] = "Sistema licenciado para GF SERVICOS DE CONTABILIDADE S/S"
        self.ws[f'A{self.current_row}'].font = Font(name='Arial', size=7)

