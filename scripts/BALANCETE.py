# Imports e Configurações Globais
import pyodbc
import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from datetime import datetime
import decimal
import json
import threading
from concurrent.futures import ThreadPoolExecutor
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from file_renamer import rename_file_after_generation

CUSTOM_PAGESIZE = (A4[0] * 1.4, A4[1] * 1.5)
CONN_STR = (
    "DRIVER={SQL Anywhere 17};"
    "HOST=NOTE-GO-273.go.local:2638;"
    "DBN=contabil;"
    "UID=ESTATISTICA002;"
    "PWD=U0T/wq6OdZ0oYSpvJRWGfg==;"
)
SUMMARY_CATEGORIES = {
    "ATIVO": {"prefixes": ["1"], "natureza": "devedora", "eh_resultado": False},
    "PASSIVO": {"prefixes": ["2"], "natureza": "credora", "eh_resultado": False},
    "RECEITAS": {"prefixes": ["3"], "natureza": "credora", "eh_resultado": True},
    "CUSTOS DAS VENDAS": {"prefixes": ["4"], "natureza": "devedora", "eh_resultado": True},
    "DESPESAS OPERACIONAIS": {"prefixes": ["5"], "natureza": "devedora", "eh_resultado": True},
    "RECEITAS LIQUIDAS DOS PRODUTOS E SERVICOS - SCP": {"prefixes": ["61"], "natureza": "credora", "eh_resultado": True},
    "DESPESAS OPERACIONAIS - SCP": {"prefixes": ["62"], "natureza": "devedora", "eh_resultado": True},
    "APURACAO DE RESULTADO - TRANSITORIA": {"prefixes": ["9"], "natureza": "transitoria", "eh_resultado": False}
}

# Dicionários de tradução
TRANSLATIONS = {
    'pt': {
        'empresa': 'Empresa',
        'periodo': 'Período', 
        'folha': 'Folha',
        'numero_livro': 'Número livro',
        'emissao': 'Emissão',
        'hora': 'Hora',
        'balancete': 'BALANCETE',
        'consolidado': 'CONSOLIDADO',
        'empresas': 'Empresas',
        'codigo': 'Código',
        'classificacao': 'Classificação', 
        'descricao': 'Descrição de Contas',
        'saldo_anterior': 'Saldo Anterior',
        'debito': 'Débito',
        'credito': 'Crédito',
        'saldo_atual': 'Saldo Atual',
        'resumo_balancete': 'RESUMO DO BALANCETE',
        'categoria': 'Categoria',
        'ativo': 'ATIVO',
        'passivo': 'PASSIVO',
        'receitas': 'RECEITAS', 
        'custos_vendas': 'CUSTOS DAS VENDAS',
        'despesas_operacionais': 'DESPESAS OPERACIONAIS',
        'receitas_liquidas_scp': 'RECEITAS LIQUIDAS DOS PRODUTOS E SERVICOS - SCP',
        'despesas_operacionais_scp': 'DESPESAS OPERACIONAIS - SCP',
        'apuracao_resultado': 'APURACAO DE RESULTADO - TRANSITORIA',
        'contas_devedoras': 'CONTAS DEVEDORAS:',
        'contas_credoras': 'CONTAS CREDORAS:',
        'resultado_mes': 'RESULTADO DO MÊS:',
        'resultado_exercicio': 'RESULTADO DO EXERCÍCIO:'
    },
    'en': {
        'empresa': 'Company',
        'periodo': 'Period',
        'folha': 'Page', 
        'numero_livro': 'Book Nº',
        'emissao': 'Date',
        'hora': 'Time',
        'balancete': 'TRIAL BALANCE SHEET',
        'consolidado': 'CONSOLIDATED',
        'empresas': 'Companies',
        'codigo': 'Code',
        'classificacao': 'Classification',
        'descricao': 'Description', 
        'saldo_anterior': 'Beginning Balance',
        'debito': 'Debit',
        'credito': 'Credit',
        'saldo_atual': 'Actual Balance',
        'resumo_balancete': 'SUMMARY OF THE TRIAL BALANCE SHEET',
        'categoria': 'Category',
        'ativo': 'ASSETS',
        'passivo': 'LIABILITIES',
        'receitas': 'REVENUES',
        'custos_vendas': 'COST OF SALES', 
        'despesas_operacionais': 'OPERATING EXPENSES',
        'receitas_liquidas_scp': 'NET REVENUES - PRODUCTS AND SERVICES - SPC',
        'despesas_operacionais_scp': 'PAYROLL COSTS - SPC',
        'apuracao_resultado': 'CALCULATION RESULT - TRANSITORY',
        'contas_devedoras': 'DEBTOR ACCOUNTS:',
        'contas_credoras': 'CREDITOR ACCOUNTS:',
        'resultado_mes': 'MONTH PROFIT/LOSS:',
        'resultado_exercicio': 'PERIOD PROFIT/LOSS:'
    }
}

def obter_parametros_relatorio_balancete(codi_emp: int):
    """
    Busca os parâmetros de configuração dinâmicos essenciais para o relatório.
    """
    print(f"\n--- Iniciando busca de parâmetros para Empresa: {codi_emp} ---")
    conn = None
    try:
        conn = pyodbc.connect(CONN_STR)
        cursor = conn.cursor()
        
        # Passo 1: Buscar a lista de filiais vinculadas à empresa principal.
        print("Passo 1: Buscando filiais para consolidação...")
        sql_empresa_principal = "SELECT cgce_emp FROM bethadba.geempre WHERE codi_emp = ?"
        cursor.execute(sql_empresa_principal, codi_emp)
        empresa_row = cursor.fetchone()
        if not empresa_row or not empresa_row[0]:
            raise ValueError(f"Empresa com código {codi_emp} não encontrada ou não possui CNPJ.")
        
        cnpj_limpo = ''.join(filter(str.isdigit, empresa_row[0]))
        radical_cnpj = cnpj_limpo[:8]
        print(f"  > Radical do CNPJ encontrado: {radical_cnpj}")
        
        lista_filiais = [codi_emp]
        sql_filiais = """
            SELECT CODI_EMP FROM BETHADBA.GEEMPRE 
            WHERE CODI_EMP <> ? AND (UCTA_EMP = 1 OR UCXA_EMP = 1) 
            AND LEFT(REPLACE(REPLACE(REPLACE(CGCE_EMP, '.', ''), '/', ''), '-', ''), 8) = ?
        """
        cursor.execute(sql_filiais, codi_emp, radical_cnpj)
        filiais_encontradas = [row.CODI_EMP for row in cursor.fetchall()]
        if filiais_encontradas:
            lista_filiais.extend(filiais_encontradas)
            print(f"  > Filiais encontradas para consolidação: {filiais_encontradas}")
        
        filiais_in_clause = ', '.join(map(str, lista_filiais))
        print(f"  > Lista final de filiais para a consulta: {lista_filiais}")

        # Passo 2: Verificando ordenação especial
        print("Passo 2: Verificando ordenação especial...")
        sql_regra = f"SELECT VALOR FROM BETHADBA.GEINICIAL WHERE CHAVE = 'Rel_Balancete' AND SECAO = 'Ordem_conta_analiticas_{codi_emp}'"
        cursor.execute(sql_regra)
        regra_row = cursor.fetchone()
        classificacoes_especiais = []
        if regra_row and regra_row[0]:
            print("  > Regra de ordenação encontrada. Buscando detalhes...")
            sql_codigos = f"SELECT VALOR FROM BETHADBA.GEINICIAL WHERE CHAVE = 'Rel_Balancete' AND SECAO = 'Ordem_conta_analiticas_contas_{codi_emp}'"
            cursor.execute(sql_codigos)
            codigos_row = cursor.fetchone()
            if codigos_row and codigos_row[0]:
                codigos_cta_list = [c.strip() for c in codigos_row[0].split(',') if c.strip()]
                for codi_cta in codigos_cta_list:
                    sql_class = "SELECT CLAS_CTA FROM BETHADBA.CTCONTAS WHERE CODI_EMP = ? AND CODI_CTA = ?"
                    cursor.execute(sql_class, codi_emp, int(codi_cta))
                    clas_row = cursor.fetchone()
                    if clas_row:
                        classificacoes_especiais.append(clas_row[0])
        
        if classificacoes_especiais:
            print(f"  > Classificações para ordenação especial: {classificacoes_especiais}")
            in_clause_items = [f"'{c}'" for c in classificacoes_especiais]
            ordem_when_condition = f"LEFT(CTCONTAS.CLAS_CTA, 6) IN ({', '.join(in_clause_items)})"
            ordem_len_param = 6
        else:
            print("  > Nenhuma regra de ordenação especial aplicada.")
            ordem_when_condition = "1 = 0"
            ordem_len_param = 0

        print("\n--- Coleta de parâmetros finalizada com sucesso! ---")
        return {
            "lista_filiais": lista_filiais,
            "filiais_in_clause": filiais_in_clause,
            "ordem_when_condition": ordem_when_condition,
            "ordem_len_param": ordem_len_param,
        }
    except pyodbc.Error as ex:
        print(f"ERRO de Banco de Dados: {ex.args[0]}\n{ex}")
        return None
    except Exception as e:
        print(f"Ocorreu um erro inesperado na coleta de parâmetros: {e}")
        return None
    finally:
        if conn:
            conn.close()

def gerar_consulta_otimizada(params, data_inicial, data_final, ingles=False):
    """
    Gera uma única consulta SQL otimizada usando CTEs para substituir tabelas temporárias.
    A consulta traz TODOS os dados - filtros de negócio são aplicados depois em Python.
    """
    empresa_principal = params['empresa_principal']
    filiais_str = params['filiais_in_clause']
    data_inicial_sql = data_inicial.replace('-', '')
    data_final_sql = data_final.replace('-', '')
    
    # CORREÇÃO FINAL: Só duplica se realmente houver múltiplas filiais
    if ',' in filiais_str:
        # Múltiplas filiais - duplicar como no original
        filiais_str_duplo = f"({filiais_str},{filiais_str})"
    else:
        # Empresa única - NÃO duplicar
        filiais_str_duplo = f"({filiais_str})"

    # Informar sobre a estratégia de filtros
    print("   ✅ Filtros rigorosos serão aplicados na camada de negócios (Python)")
    print("   📊 Consulta SQL traz TODOS os dados - filtros aplicados depois")

    # Definir campo NOME_CTA baseado no idioma
    if ingles:
        nome_cta_field = "COALESCE(CTCONTAS_IDIOMAS.DESCRICAO, CTCONTAS.NOME_CTA) AS NOME_CTA"
        join_idiomas = "LEFT OUTER JOIN BETHADBA.CTCONTAS_IDIOMAS ON (CTCONTAS_IDIOMAS.CODI_EMP = CTCONTAS.CODI_EMP AND CTCONTAS_IDIOMAS.CODI_CTA = CTCONTAS.CODI_CTA AND CTCONTAS_IDIOMAS.I_IDIOMAS = 1)"
        print("   🌐 Modo inglês ativado - usando traduções da tabela CTCONTAS_IDIOMAS")
    else:
        nome_cta_field = "CTCONTAS.NOME_CTA"
        join_idiomas = ""
        print("   🇧🇷 Modo português - usando nomes originais das contas")

    # Consulta única consolidada com CTEs
    consulta_otimizada = f"""
        WITH 
        saldo_ant_debit AS (
            SELECT LANX.CODI_EMP, LANX.CDEB_LAN, LANX.FILI_LAN, MAX(LANX.DATA_LAN) AS DATA_LAN, 
                   COALESCE(SUM(ROUND(LANX.VLOR_LAN / COALESCE(TDINDICE.VALOR_INDICE, 1), 6)), 0) AS VLOR_LAN
            FROM BETHADBA.CTLANCTO LANX, 
                 LATERAL(SELECT (DSDBA.FG_BUSCA_INDICE_CT('N', 'CORRENTE', {data_final_sql}, LANX.DATA_LAN)) AS VALOR_INDICE FROM DSDBA.DUMMY) AS TDINDICE
            WHERE LANX.fili_lan IN {filiais_str_duplo}
              AND LANX.CODI_EMP_PLANO = {empresa_principal}
              AND LANX.DATA_LAN_BUSCA < {data_inicial_sql}
              AND LANX.CDEB_LAN > 0 
            GROUP BY LANX.CODI_EMP, LANX.CDEB_LAN, LANX.FILI_LAN
        ),
        saldo_ant_credit AS (
            SELECT LANX.CODI_EMP, LANX.CCRE_LAN, LANX.FILI_LAN, MAX(LANX.DATA_LAN) AS DATA_LAN, 
                   COALESCE(SUM(ROUND(LANX.VLOR_LAN / COALESCE(TDINDICE.VALOR_INDICE, 1), 6)), 0) AS VLOR_LAN
            FROM BETHADBA.CTLANCTO LANX, 
                 LATERAL(SELECT (DSDBA.FG_BUSCA_INDICE_CT('N', 'CORRENTE', {data_final_sql}, LANX.DATA_LAN)) AS VALOR_INDICE FROM DSDBA.DUMMY) AS TDINDICE
            WHERE LANX.fili_lan IN {filiais_str_duplo}
              AND LANX.CODI_EMP_PLANO = {empresa_principal}
              AND LANX.DATA_LAN_BUSCA < {data_inicial_sql}
              AND LANX.CCRE_LAN > 0 
            GROUP BY LANX.CODI_EMP, LANX.CCRE_LAN, LANX.FILI_LAN
        ),
        mov_atual_debit AS (
            SELECT LANX.CODI_EMP, LANX.CDEB_LAN, LANX.FILI_LAN, MAX(LANX.DATA_LAN) AS DATA_LAN, 
                   COALESCE(SUM(ROUND(LANX.VLOR_LAN / COALESCE(TDINDICE.VALOR_INDICE, 1), 6)), 0) AS VLOR_LAN
            FROM BETHADBA.CTLANCTO LANX, 
                 LATERAL(SELECT (DSDBA.FG_BUSCA_INDICE_CT('N', 'CORRENTE', {data_final_sql}, LANX.DATA_LAN)) AS VALOR_INDICE FROM DSDBA.DUMMY) AS TDINDICE
            WHERE LANX.fili_lan IN {filiais_str_duplo}
              AND LANX.CODI_EMP_PLANO = {empresa_principal}
              AND LANX.DATA_LAN_BUSCA BETWEEN {data_inicial_sql} AND {data_final_sql}
              AND LANX.CDEB_LAN > 0 
            GROUP BY LANX.CODI_EMP, LANX.CDEB_LAN, LANX.FILI_LAN
        ),
        mov_atual_credit AS (
            SELECT LANX.CODI_EMP, LANX.CCRE_LAN, LANX.FILI_LAN, MAX(LANX.DATA_LAN) AS DATA_LAN, 
                   COALESCE(SUM(ROUND(LANX.VLOR_LAN / COALESCE(TDINDICE.VALOR_INDICE, 1), 6)), 0) AS VLOR_LAN
            FROM BETHADBA.CTLANCTO LANX, 
                 LATERAL(SELECT (DSDBA.FG_BUSCA_INDICE_CT('N', 'CORRENTE', {data_final_sql}, LANX.DATA_LAN)) AS VALOR_INDICE FROM DSDBA.DUMMY) AS TDINDICE
            WHERE LANX.fili_lan IN {filiais_str_duplo}
              AND LANX.CODI_EMP_PLANO = {empresa_principal}
              AND LANX.DATA_LAN_BUSCA BETWEEN {data_inicial_sql} AND {data_final_sql}
              AND LANX.CCRE_LAN > 0 
            GROUP BY LANX.CODI_EMP, LANX.CCRE_LAN, LANX.FILI_LAN
        )
        SELECT CTCONTAS.CODI_CTA, {nome_cta_field}, CTCONTAS.CLAS_CTA, CTCONTAS.TIPO_CTA, 
            TDSALDO_ANTERIOR.SALDO AS SALDOANT, TDTOTAL_DEBITO.VALOR_TOTAL AS TOTDEB, 
            TDTOTAL_CREDITO.VALOR_TOTAL AS TOTCRE, TDSALDO_ATUAL.SALDO AS SALDOATU,
            GEEMPRE.CGCE_EMP, CURRENT DATE AS EMISSAO_DATA, CURRENT TIME AS EMISSAO_HORA,
            DATE('{data_inicial}') AS DATINI, DATE('{data_final}') AS DATFIN, TD_ORDEM.ORDEM AS ORDEM
        FROM BETHADBA.CTCONTAS AS CTCONTAS
        {join_idiomas}
        INNER JOIN BETHADBA.CTPARMTO ON CTPARMTO.CODI_EMP = {empresa_principal}
        INNER JOIN BETHADBA.GEEMPRE ON GEEMPRE.CODI_EMP = CTPARMTO.CODI_EMP,
        LATERAL(SELECT COALESCE(SUM(LANX.VLOR_LAN), 0) FROM saldo_ant_debit LANX WHERE LANX.CODI_EMP = {empresa_principal} AND LANX.CDEB_LAN = CTCONTAS.CODI_CTA) AS TDTOTAL_DEBITO_ANTERIOR(VALOR_TOTAL),
        LATERAL(SELECT COALESCE(SUM(LANX.VLOR_LAN), 0) FROM saldo_ant_credit LANX WHERE LANX.CODI_EMP = {empresa_principal} AND LANX.CCRE_LAN = CTCONTAS.CODI_CTA) AS TDTOTAL_CREDITO_ANTERIOR(VALOR_TOTAL),
        LATERAL(SELECT COALESCE(SUM(LANX.VLOR_LAN), 0) FROM mov_atual_debit LANX WHERE LANX.CODI_EMP = {empresa_principal} AND LANX.CDEB_LAN = CTCONTAS.CODI_CTA) AS TDTOTAL_DEBITO(VALOR_TOTAL),
        LATERAL(SELECT COALESCE(SUM(LANX.VLOR_LAN), 0) FROM mov_atual_credit LANX WHERE LANX.CODI_EMP = {empresa_principal} AND LANX.CCRE_LAN = CTCONTAS.CODI_CTA) AS TDTOTAL_CREDITO(VALOR_TOTAL),
        LATERAL(SELECT CASE WHEN CTCONTAS.TIPO_CTA = 'A' THEN TDTOTAL_DEBITO_ANTERIOR.VALOR_TOTAL - TDTOTAL_CREDITO_ANTERIOR.VALOR_TOTAL ELSE 0 END FROM DSDBA.DUMMY) AS TDSALDO_ANTERIOR(SALDO),
        LATERAL(SELECT CASE WHEN CTCONTAS.TIPO_CTA = 'A' THEN TDSALDO_ANTERIOR.SALDO + TDTOTAL_DEBITO.VALOR_TOTAL - TDTOTAL_CREDITO.VALOR_TOTAL ELSE 0 END FROM DSDBA.DUMMY) AS TDSALDO_ATUAL(SALDO),
        LATERAL(SELECT CASE WHEN {params['ordem_len_param']} > 0 AND CTCONTAS.TIPO_CTA = 'A' AND {params['ordem_when_condition']} THEN LEFT(CTCONTAS.CLAS_CTA, {params['ordem_len_param']}) || '001' ELSE CTCONTAS.CLAS_CTA END FROM DSDBA.DUMMY) AS TD_ORDEM(ORDEM)
        WHERE CTCONTAS.CODI_EMP = {empresa_principal}
        AND TRIM(CTCONTAS.CLAS_CTA) <> '' 
        ORDER BY ORDEM ASC, NOME_CTA ASC
    """
    
    return consulta_otimizada

def aplicar_filtros_negocios(df):
    """
    Aplica filtros de regras de negócio no DataFrame após a consulta SQL.
    Separação de responsabilidades: SQL traz dados, Python aplica regras.
    """
    print("🔍 Aplicando filtros de regras de negócio...")
    
    df_original = df.copy()
    linhas_iniciais = len(df_original)
    
    # ========== FILTROS BÁSICOS (SEMPRE APLICADOS) ==========
    print("   🔧 Aplicando filtros básicos...")
    
    # 1. REMOVER APENAS CONTAS DE PREJUÍZO (apenas 238 - NÃO 230!)
    linhas_antes = len(df)
    # Remover apenas contas que começam EXATAMENTE com 238 (Prejuízo)
    mask_prejuizo = df['CLAS_CTA'].astype(str).str.match(r'^238')
    df = df[~mask_prejuizo]
    removidas_prejuizo = linhas_antes - len(df)
    if removidas_prejuizo > 0:
        print(f"      ❌ {removidas_prejuizo} contas de Prejuízo (238) removidas")
    print(f"      ✅ Contas 230 (Lucro) e 232 preservadas - apenas 238 (Prejuízo) removida")
    
    # 2. PARA ANALÍTICAS: Remover completamente zeradas
    linhas_antes = len(df)
    mask_analiticas_zeradas = (
        (df['TIPO_CTA'] == 'A') & 
        (df['SALDOANT'] == 0) & 
        (df['TOTDEB'] == 0) & 
        (df['TOTCRE'] == 0) & 
        (df['SALDOATU'] == 0)
    )
    df = df[~mask_analiticas_zeradas]
    removidas_zeradas = linhas_antes - len(df)
    if removidas_zeradas > 0:
        print(f"      ❌ {removidas_zeradas} contas analíticas completamente zeradas removidas")
    
    # 3. PARA ANALÍTICAS: Exigir TOTDEB >= 0 E TOTCRE >= 0
    linhas_antes = len(df)
    mask_sem_movimento = (
        (df['TIPO_CTA'] == 'A') & 
        ((df['TOTDEB'] < 0) | (df['TOTCRE'] < 0))
    )
    df = df[~mask_sem_movimento]
    removidas_sem_movimento = linhas_antes - len(df)
    if removidas_sem_movimento > 0:
        print(f"      ❌ {removidas_sem_movimento} contas analíticas sem movimento (TOTDEB < 0 ou TOTCRE < 0) removidas")
    
    # ========== FILTROS RIGOROSOS (SEMPRE APLICADOS) ==========
    print("   🔧 Aplicando filtros rigorosos...")
    
    # 4. REMOVER ANALÍTICAS com saldos anterior/atual = 0
    linhas_antes = len(df)
    mask_saldos_zerados = (
        (df['TIPO_CTA'] == 'A') & 
        (df['SALDOANT'] == 0) & 
        (df['SALDOATU'] == 0)
    )
    df = df[~mask_saldos_zerados]
    removidas_saldos = linhas_antes - len(df)
    if removidas_saldos > 0:
        print(f"      ❌ {removidas_saldos} contas analíticas com saldos zerados removidas")
    
    # ========== RESUMO ==========
    linhas_finais = len(df)
    sinteticas_finais = len(df[df['TIPO_CTA'] == 'S'])
    analiticas_finais = len(df[df['TIPO_CTA'] == 'A'])
    
    print(f"   📊 Resultado: {linhas_iniciais} → {linhas_finais} linhas ({sinteticas_finais} sintéticas + {analiticas_finais} analíticas)")
    print(f"   ✅ {linhas_iniciais - linhas_finais} linhas removidas pelos filtros")
    
    return df

def buscar_dados_auxiliares(empresa_principal):
    """
    Busca dados de cabeçalho e rodapé em paralelo.
    """
    conn = None
    try:
        conn = pyodbc.connect(CONN_STR)
        cursor = conn.cursor()
        
        # Buscar informações de cabeçalho
        header_query = f"SELECT RAZAO_EMP, CGCE_EMP FROM BETHADBA.GEEMPRE WHERE CODI_EMP = {empresa_principal}"
        cursor.execute(header_query)
        header_row = cursor.fetchone()
        
        # Buscar representante legal
        legal_query = f"SELECT RLEG_EMP, CPF_LEG_EMP FROM BETHADBA.GEEMPRE WHERE CODI_EMP = {empresa_principal}"
        cursor.execute(legal_query)
        rep_legal_row = cursor.fetchone()
        
        # Buscar contador
        contador_query = "SELECT NOME_CON, RCRC_CON, CPFC_CON, UF_CRC FROM BETHADBA.GECONTADOR WHERE CODI_CON = 5"
        cursor.execute(contador_query)
        contador_row = cursor.fetchone()
        
        header_data = {"razao_emp": header_row.RAZAO_EMP, "cnpj": header_row.CGCE_EMP}
        footer_data = {
            "rep_legal_nome": rep_legal_row.RLEG_EMP, "rep_legal_cpf": rep_legal_row.CPF_LEG_EMP,
            "contador_nome": contador_row.NOME_CON, "contador_crc": contador_row.RCRC_CON,
            "contador_cpf": contador_row.CPFC_CON, "contador_uf_crc": contador_row.UF_CRC,
        }
        
        return header_data, footer_data
        
    except Exception as e:
        print(f"Erro ao buscar dados auxiliares: {e}")
        return None, None
    finally:
        if conn:
            conn.close()

# Funções de formatação
def format_cnpj(cnpj):
    cnpj = ''.join(filter(str.isdigit, str(cnpj)))
    if len(cnpj) != 14: return cnpj
    return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"

def format_classificacao(clas_str):
    s = str(clas_str).strip()
    if not s: return ""
    parts = []
    if len(s) >= 1: parts.append(s[0])
    if len(s) >= 2: parts.append(s[1])
    if len(s) >= 3: parts.append(s[2])
    if len(s) >= 4: parts.append(s[3:6])
    if len(s) >= 7: parts.append(s[6:])
    return ".".join(filter(None, parts))

def format_currency(value, show_suffix=True):
    if value is None or not isinstance(value, (int, float, decimal.Decimal)): 
        value = decimal.Decimal('0.00')
    else: 
        value = decimal.Decimal(value)
    if abs(value) < decimal.Decimal('0.005'): 
        return "0,00"
    formatted_value = f"{abs(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    if not show_suffix: 
        return formatted_value
    suffix = 'D' if value >= 0 else 'C'
    return f"{formatted_value}{suffix}"

def format_currency_resultado(value):
    if value is None or not isinstance(value, (int, float, decimal.Decimal)): 
        value = decimal.Decimal('0.00')
    else: 
        value = decimal.Decimal(value)
    if abs(value) < decimal.Decimal('0.005'): 
        return "0,00"
    formatted_value = f"{abs(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    suffix = 'C' if value >= 0 else 'D'
    return f"{formatted_value}{suffix}"

def calcular_nivel_hierarquico(classificacao):
    s = str(classificacao).strip()
    if len(s) == 1: return 1
    elif len(s) == 2: return 2
    elif len(s) == 3: return 3
    elif len(s) <= 6: return 4
    else: return 5

def format_nome_hierarquico(nome, nivel):
    return "  " * (nivel - 1) + nome

def calcular_totalizacao_hierarquica_otimizada(df):
    """
    Versão otimizada do cálculo hierárquico usando operações vetorizadas do pandas.
    """
    print("   Calculando totalização hierárquica (versão otimizada)...")
    df_calc = df.copy()
    contas_sinteticas = df_calc[df_calc['TIPO_CTA'] == 'S'].copy()
    contas_analiticas = df_calc[df_calc['TIPO_CTA'] == 'A'].copy()

    print(f"   DEBUG: Total de contas sintéticas: {len(contas_sinteticas)}")
    print(f"   DEBUG: Total de contas analíticas: {len(contas_analiticas)}")

    # Criar um mapeamento de classificação para otimizar a busca
    analiticas_dict = {}
    for idx, row in contas_analiticas.iterrows():
        clas = str(row['CLAS_CTA']).strip()
        analiticas_dict[clas] = {
            'SALDOANT': row['SALDOANT'],
            'TOTDEB': row['TOTDEB'], 
            'TOTCRE': row['TOTCRE'],
            'SALDOATU': row['SALDOATU']
        }

    # Processar sintéticas de forma otimizada
    for idx, conta_sintetica in contas_sinteticas.iterrows():
        clas_sintetica = str(conta_sintetica['CLAS_CTA']).strip()
        
        # Buscar descendentes usando list comprehension (mais rápido)
        descendentes_dados = []
        for clas, dados in analiticas_dict.items():
            if any(clas.startswith(clas_sintetica + str(i)) for i in range(10)):
                descendentes_dados.append(dados)
        
        # DEBUG apenas para contas que não sejam de Lucro e Prejuízo (230, 238)
        if clas_sintetica in ['11', '12', '21', '22'] and not clas_sintetica.startswith(('230', '238')):
            print(f"   DEBUG: '{clas_sintetica}' -> {len(descendentes_dados)} descendentes")
        
        if descendentes_dados:
            # Somar usando pandas (operação vetorizada)
            df_calc.at[idx, 'SALDOANT'] = sum(d['SALDOANT'] for d in descendentes_dados)
            df_calc.at[idx, 'TOTDEB'] = sum(d['TOTDEB'] for d in descendentes_dados)
            df_calc.at[idx, 'TOTCRE'] = sum(d['TOTCRE'] for d in descendentes_dados)
            df_calc.at[idx, 'SALDOATU'] = sum(d['SALDOATU'] for d in descendentes_dados)
            
            # DEBUG apenas para contas que não sejam de Lucro e Prejuízo (230, 238)
            if clas_sintetica in ['11', '12', '21', '22'] and not clas_sintetica.startswith(('230', '238')):
                print(f"   DEBUG: '{clas_sintetica}' TOTAL = {df_calc.at[idx, 'SALDOATU']}")
    
    return df_calc.sort_values(['ORDEM', 'NOME_CTA'])

def draw_header(c, width, page_num, header_data, params, data_inicial, data_final, ingles=False):
    lang = 'en' if ingles else 'pt'
    t = TRANSLATIONS[lang]
    
    c.setFont("Helvetica-Bold", 10)
    c.drawString(1.5 * cm, CUSTOM_PAGESIZE[1] - 2 * cm, f"{t['empresa']}: {header_data['razao_emp']}")
    formatted_cnpj = format_cnpj(header_data['cnpj'])
    c.drawString(1.5 * cm, CUSTOM_PAGESIZE[1] - 2.5 * cm, f"C.N.P.J.: {formatted_cnpj}")
    c.drawString(1.5 * cm, CUSTOM_PAGESIZE[1] - 3 * cm, f"{t['periodo']}: {datetime.strptime(data_inicial, '%Y-%m-%d').strftime('%d/%m/%Y')} - {datetime.strptime(data_final, '%Y-%m-%d').strftime('%d/%m/%Y')}")
    
    # Só mostrar CONSOLIDADO se houver múltiplas empresas
    filiais_tuple = params['lista_filiais']
    y_consolidado = CUSTOM_PAGESIZE[1] - 3.5 * cm
    if len(filiais_tuple) > 1:
        c.drawString(1.5 * cm, y_consolidado, f"{t['consolidado']} ({t['empresas']}: {', '.join(map(str, filiais_tuple))})")
        y_titulo = CUSTOM_PAGESIZE[1] - 4.5 * cm
        y_cabecalho = CUSTOM_PAGESIZE[1] - 5.5 * cm
        y_inicio_dados = CUSTOM_PAGESIZE[1] - 6 * cm
    else:
        # Se não há consolidado, move tudo para cima
        y_titulo = CUSTOM_PAGESIZE[1] - 4 * cm
        y_cabecalho = CUSTOM_PAGESIZE[1] - 5 * cm
        y_inicio_dados = CUSTOM_PAGESIZE[1] - 5.5 * cm
    
    c.setFont("Helvetica", 9)
    c.drawRightString(width - 1.5 * cm, CUSTOM_PAGESIZE[1] - 2 * cm, f"{t['folha']}: {page_num:04d}")
    c.drawRightString(width - 1.5 * cm, CUSTOM_PAGESIZE[1] - 2.5 * cm, f"{t['numero_livro']}: 0001")
    c.drawRightString(width - 1.5 * cm, CUSTOM_PAGESIZE[1] - 3 * cm, f"{t['emissao']}: {datetime.now().strftime('%d/%m/%Y')}")
    c.drawRightString(width - 1.5 * cm, y_consolidado, f"{t['hora']}: {datetime.now().strftime('%H:%M:%S')}")
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width / 2.0, y_titulo, t['balancete'])
    c.setFont("Helvetica-Bold", 8)
    c.drawRightString(2.8 * cm, y_cabecalho, t['codigo'])
    c.drawString(3 * cm, y_cabecalho, t['classificacao'])
    c.drawString(5.5 * cm, y_cabecalho, t['descricao'])
    c.drawRightString(19 * cm, y_cabecalho, t['saldo_anterior'])
    c.drawRightString(22 * cm, y_cabecalho, t['debito'])
    c.drawRightString(25 * cm, y_cabecalho, t['credito'])
    c.drawRightString(28 * cm, y_cabecalho, t['saldo_atual'])
    c.line(1.5 * cm, y_cabecalho - 0.2 * cm, width - 1.5 * cm, y_cabecalho - 0.2 * cm)
    
    return y_inicio_dados

def draw_summary_section(c, df_completo, y_pos, width, height, page_num, header_data, natureza_contas_9, params, data_inicial, data_final, ingles=False):
    """
    VERSÃO FINAL CORRIGIDA - Usa apenas contas analíticas para calcular totais por categoria
    """
    lang = 'en' if ingles else 'pt'
    t = TRANSLATIONS[lang]
    
    print("   Iniciando desenho da seção de resumo...")
    needed_height = 12 * cm
    
    if y_pos < needed_height:
        c.showPage()
        page_num += 1
        y_pos = draw_header(c, width, page_num, header_data, params, data_inicial, data_final, ingles)
    
    # Usar apenas contas analíticas para evitar duplicação na totalização
    df_analiticas = df_completo[df_completo['TIPO_CTA'] == 'A'].copy()
    
    y_pos -= 1.5 * cm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(1.5 * cm, y_pos, t['resumo_balancete'])
    y_pos -= 0.7 * cm
    
    c.setFont("Helvetica-Bold", 8)
    c.drawString(1.5 * cm, y_pos, t['categoria'])
    c.drawRightString(19 * cm, y_pos, t['saldo_anterior'])
    c.drawRightString(22 * cm, y_pos, t['debito'])
    c.drawRightString(25 * cm, y_pos, t['credito'])
    c.drawRightString(28 * cm, y_pos, t['saldo_atual'])
    y_pos -= 0.2 * cm
    c.line(1.5 * cm, y_pos, width - 3 * cm, y_pos)
    y_pos -= 0.5 * cm
    
    c.setFont("Helvetica", 8)
    
    # Mapeamento de categorias para traduções
    category_translations = {
        "ATIVO": t['ativo'],
        "PASSIVO": t['passivo'], 
        "RECEITAS": t['receitas'],
        "CUSTOS DAS VENDAS": t['custos_vendas'],
        "DESPESAS OPERACIONAIS": t['despesas_operacionais'],
        "RECEITAS LIQUIDAS DOS PRODUTOS E SERVICOS - SCP": t['receitas_liquidas_scp'],
        "DESPESAS OPERACIONAIS - SCP": t['despesas_operacionais_scp'],
        "APURACAO DE RESULTADO - TRANSITORIA": t['apuracao_resultado']
    }
    
    # Calcular totais por categoria usando APENAS contas analíticas
    summary_data = {}
    for nome, config in SUMMARY_CATEGORIES.items():
        if config['prefixes'] in [["61"], ["62"], ["9"]]:
            # SCP e Transitória ficam zero
            summary_data[nome] = {'saldo_ant': 0, 'tot_deb': 0, 'tot_cre': 0, 'saldo_atu': 0}
        else:
            # Filtrar contas analíticas que começam com os prefixos da categoria
            mask = df_analiticas['CLAS_CTA'].str.startswith(tuple(config['prefixes']))
            df_grupo = df_analiticas[mask]
            
            summary_data[nome] = {
                'saldo_ant': df_grupo['SALDOANT'].sum(),
                'tot_deb': df_grupo['TOTDEB'].sum(),
                'tot_cre': df_grupo['TOTCRE'].sum(),
                'saldo_atu': df_grupo['SALDOATU'].sum()
            }
        
        # Usar tradução da categoria
        categoria_nome = category_translations.get(nome, nome)
        c.drawString(1.5 * cm, y_pos, categoria_nome)
        c.drawRightString(19 * cm, y_pos, format_currency(summary_data[nome]['saldo_ant']))
        c.drawRightString(22 * cm, y_pos, format_currency(summary_data[nome]['tot_deb'], show_suffix=False))
        c.drawRightString(25 * cm, y_pos, format_currency(summary_data[nome]['tot_cre'], show_suffix=False))
        c.drawRightString(28 * cm, y_pos, format_currency(summary_data[nome]['saldo_atu']))
        y_pos -= 0.5 * cm
        
        # Duplicar a linha "APURACAO DE RESULTADO - TRANSITORIA"
        if nome == "APURACAO DE RESULTADO - TRANSITORIA":
            c.drawString(1.5 * cm, y_pos, categoria_nome)
            c.drawRightString(19 * cm, y_pos, format_currency(summary_data[nome]['saldo_ant']))
            c.drawRightString(22 * cm, y_pos, format_currency(summary_data[nome]['tot_deb'], show_suffix=False))
            c.drawRightString(25 * cm, y_pos, format_currency(summary_data[nome]['tot_cre'], show_suffix=False))
            c.drawRightString(28 * cm, y_pos, format_currency(summary_data[nome]['saldo_atu']))
            y_pos -= 0.5 * cm

    y_pos -= 0.5 * cm
    
    # Totalização Devedoras e Credoras
    devedoras = {
        'saldo_ant': (summary_data['ATIVO']['saldo_ant'] + 
                      summary_data['CUSTOS DAS VENDAS']['saldo_ant'] + 
                      summary_data['DESPESAS OPERACIONAIS']['saldo_ant']),
        'tot_deb':   (summary_data['ATIVO']['tot_deb'] + 
                      summary_data['CUSTOS DAS VENDAS']['tot_deb'] + 
                      summary_data['DESPESAS OPERACIONAIS']['tot_deb']),
        'tot_cre':   (summary_data['ATIVO']['tot_cre'] + 
                      summary_data['CUSTOS DAS VENDAS']['tot_cre'] + 
                      summary_data['DESPESAS OPERACIONAIS']['tot_cre']),
        'saldo_atu': (summary_data['ATIVO']['saldo_atu'] + 
                      summary_data['CUSTOS DAS VENDAS']['saldo_atu'] + 
                      summary_data['DESPESAS OPERACIONAIS']['saldo_atu'])
    }
    credoras = {
        'saldo_ant': summary_data['PASSIVO']['saldo_ant'] + summary_data['RECEITAS']['saldo_ant'],
        'tot_deb':   summary_data['PASSIVO']['tot_deb'] + summary_data['RECEITAS']['tot_deb'],
        'tot_cre':   summary_data['PASSIVO']['tot_cre'] + summary_data['RECEITAS']['tot_cre'],
        'saldo_atu': summary_data['PASSIVO']['saldo_atu'] + summary_data['RECEITAS']['saldo_atu']
    }
    
    # Apuração do Resultado usando analíticas
    receitas_analiticas = df_analiticas[df_analiticas['CLAS_CTA'].str.startswith('3')]
    custos_analiticas = df_analiticas[df_analiticas['CLAS_CTA'].str.startswith('4')]
    despesas_analiticas = df_analiticas[df_analiticas['CLAS_CTA'].str.startswith('5')]
    
    debito_total_mes = (custos_analiticas['TOTDEB'].sum() - custos_analiticas['TOTCRE'].sum()) + (despesas_analiticas['TOTDEB'].sum() - despesas_analiticas['TOTCRE'].sum())
    credito_total_mes = receitas_analiticas['TOTCRE'].sum() - receitas_analiticas['TOTDEB'].sum()
    resultado_mes = credito_total_mes - debito_total_mes

    saldo_anterior_exercicio = (custos_analiticas['SALDOANT'].sum() + despesas_analiticas['SALDOANT'].sum()) + receitas_analiticas['SALDOANT'].sum()
    debito_exercicio = custos_analiticas['SALDOATU'].sum() + despesas_analiticas['SALDOATU'].sum()
    credito_exercicio = abs(receitas_analiticas['SALDOATU'].sum())
    resultado_exercicio = credito_exercicio - debito_exercicio
    
    # Exibição no PDF
    c.setFont("Helvetica-Bold", 8)
    c.drawString(1.5 * cm, y_pos, t['contas_devedoras'])
    c.drawRightString(19 * cm, y_pos, format_currency(devedoras['saldo_ant']))
    c.drawRightString(22 * cm, y_pos, format_currency(devedoras['tot_deb'], show_suffix=False))
    c.drawRightString(25 * cm, y_pos, format_currency(devedoras['tot_cre'], show_suffix=False))
    c.drawRightString(28 * cm, y_pos, format_currency(devedoras['saldo_atu']))
    y_pos -= 0.5 * cm

    c.drawString(1.5 * cm, y_pos, t['contas_credoras'])
    c.drawRightString(19 * cm, y_pos, format_currency(credoras['saldo_ant']))
    c.drawRightString(22 * cm, y_pos, format_currency(credoras['tot_deb'], show_suffix=False))
    c.drawRightString(25 * cm, y_pos, format_currency(credoras['tot_cre'], show_suffix=False))
    c.drawRightString(28 * cm, y_pos, format_currency(credoras['saldo_atu']))
    y_pos -= 0.5 * cm

    c.drawString(1.5 * cm, y_pos, t['resultado_mes'])
    c.drawString(19 * cm - 3 * cm, y_pos, "-")
    c.drawRightString(22 * cm, y_pos, format_currency(debito_total_mes, show_suffix=False))
    c.drawRightString(25 * cm, y_pos, format_currency(credito_total_mes, show_suffix=False))
    c.drawRightString(28 * cm, y_pos, format_currency_resultado(resultado_mes))
    y_pos -= 0.5 * cm

    c.drawString(1.5 * cm, y_pos, t['resultado_exercicio'])
    c.drawRightString(19 * cm, y_pos, format_currency(saldo_anterior_exercicio))
    c.drawRightString(22 * cm, y_pos, format_currency(debito_exercicio, show_suffix=False))
    c.drawRightString(25 * cm, y_pos, format_currency(credito_exercicio, show_suffix=False))
    c.drawRightString(28 * cm, y_pos, format_currency_resultado(resultado_exercicio))
    
    return y_pos, page_num

def draw_footer(c, width, footer_data):
    y = 3 * cm
    c.setFont("Helvetica", 8)
    c.drawCentredString(width / 4, y, "_" * 40)
    c.drawCentredString(width * 3 / 4, y, "_" * 40)
    y -= 0.5 * cm
    c.drawCentredString(width / 4, y, footer_data['rep_legal_nome'])
    c.drawCentredString(width * 3 / 4, y, footer_data['contador_nome'])
    y -= 0.5 * cm
    c.drawCentredString(width / 4, y, f"CPF: {footer_data['rep_legal_cpf']}")
    c.drawCentredString(width * 3 / 4, y, f"Reg. no CRC - {footer_data['contador_uf_crc']} sob o No. {footer_data['contador_crc']}")
    y -= 0.5 * cm
    c.drawCentredString(width * 3 / 4, y, f"CPF: {footer_data['contador_cpf']}")

def generate_pdf(df_report, header_data, footer_data, params, data_inicial, data_final, timestamp_str, ingles=False):
    """
    Gera o PDF do balancete e retorna o nome do arquivo.
    """
    filename = f"Balancete_{params['empresa_principal']}_{data_inicial}_a_{data_final}_{timestamp_str}.pdf"
    c = canvas.Canvas(filename, pagesize=CUSTOM_PAGESIZE)
    width, height = CUSTOM_PAGESIZE
    
    page_num = 1
    y_pos = draw_header(c, width, page_num, header_data, params, data_inicial, data_final, ingles)
    
    print(f"   Gerando PDF com {len(df_report)} linhas...")

    linha_num = 1
    natureza_contas_9 = pd.DataFrame()

    for index, row in df_report.iterrows():
        clas_atual = str(row['CLAS_CTA'])
        nivel_atual = calcular_nivel_hierarquico(clas_atual)

        if y_pos < 5 * cm:
            c.showPage()
            page_num += 1
            y_pos = draw_header(c, width, page_num, header_data, params, data_inicial, data_final, ingles)
            linha_num = 1
        
        if linha_num % 2 == 0:
            c.setFillColorRGB(0.95, 0.95, 0.95)
            c.rect(1.5 * cm, y_pos - 0.1 * cm, width - 3 * cm, 0.5 * cm, fill=1, stroke=0)
            c.setFillColorRGB(0, 0, 0)
        
        font_name = "Helvetica-Bold" if row['TIPO_CTA'] == 'S' else "Helvetica"
        c.setFont(font_name, 8)
        
        c.drawRightString(2.8 * cm, y_pos, str(row['CODI_CTA']))
        c.drawString(3 * cm, y_pos, format_classificacao(row['CLAS_CTA']))
        
        nome_hierarquico = format_nome_hierarquico(str(row['NOME_CTA']), nivel_atual)
        # Truncar para 60 caracteres no PDF
        nome_truncado = nome_hierarquico[:60] + "..." if len(nome_hierarquico) > 60 else nome_hierarquico
        c.drawString(5.5 * cm, y_pos, nome_truncado)
        
        c.drawRightString(19 * cm, y_pos, format_currency(row['SALDOANT']))
        c.drawRightString(22 * cm, y_pos, format_currency(row['TOTDEB'], show_suffix=False))
        c.drawRightString(25 * cm, y_pos, format_currency(row['TOTCRE'], show_suffix=False))
        c.drawRightString(28 * cm, y_pos, format_currency(row['SALDOATU']))
        
        y_pos -= 0.5 * cm
        linha_num += 1
    
    y_pos, page_num = draw_summary_section(
        c, df_report, y_pos, width, height, page_num, 
        header_data, natureza_contas_9, params, data_inicial, data_final, ingles
    )
    
    if y_pos < 6 * cm:
        c.showPage()
        page_num += 1
    
    draw_footer(c, width, footer_data)
    c.save()
    print(f"\nPDF '{filename}' gerado com sucesso!")
    return filename

def generate_xlsx(df_report, header_data, footer_data, params, data_inicial, data_final, timestamp_str, ingles=False):
    """
    Gera arquivo Excel com layout idêntico ao PDF e retorna o nome do arquivo.
    """
    lang = 'en' if ingles else 'pt'
    t = TRANSLATIONS[lang]
    
    filename = f"Balancete_{params['empresa_principal']}_{data_inicial}_a_{data_final}_{timestamp_str}.xlsx"
    print(f"   Gerando XLSX '{filename}'...")
    
    # Criar workbook e worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balancete"
    
    # Remover linhas de grade completamente
    ws.sheet_view.showGridLines = False
    ws.print_options.gridLines = False
    ws.print_options.gridLinesSet = True
    
    # Configurações de página limpas
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False
    ws.page_setup.fitToWidth = False
    
    # Garantir que todas as células tenham fundo branco
    from openpyxl.styles import PatternFill
    white_fill = PatternFill(fill_type=None)  # Sem preenchimento = branco
    
    # Definir estilos
    header_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=8)
    bold_font = Font(name='Arial', size=8, bold=True)
    
    # Alinhamentos
    center_align = Alignment(horizontal='center')
    right_align = Alignment(horizontal='right')
    left_align = Alignment(horizontal='left')
    
    row = 1
    
    # CABEÇALHO DA EMPRESA - Com fundo branco garantido
    cell = ws[f'A{row}']
    cell.value = f"{t['empresa']}: {header_data['razao_emp']}"
    cell.font = header_font
    cell.fill = white_fill
    cell = ws[f'G{row}']
    cell.value = f"{t['folha']}: 0001"
    cell.font = normal_font
    cell.alignment = right_align
    cell.fill = white_fill
    row += 1
    
    cell = ws[f'A{row}']
    cell.value = f"C.N.P.J.: {format_cnpj(header_data['cnpj'])}"
    cell.font = header_font
    cell.fill = white_fill
    cell = ws[f'G{row}']
    cell.value = f"{t['numero_livro']}: 0001"
    cell.font = normal_font
    cell.alignment = right_align
    cell.fill = white_fill
    row += 1
    
    cell = ws[f'A{row}']
    cell.value = f"{t['periodo']}: {datetime.strptime(data_inicial, '%Y-%m-%d').strftime('%d/%m/%Y')} - {datetime.strptime(data_final, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    cell.font = header_font
    cell.fill = white_fill
    cell = ws[f'G{row}']
    cell.value = f"{t['emissao']}: {datetime.now().strftime('%d/%m/%Y')}"
    cell.font = normal_font
    cell.alignment = right_align
    cell.fill = white_fill
    row += 1
    
    # Só mostrar CONSOLIDADO se houver múltiplas empresas
    if len(params['lista_filiais']) > 1:
        cell = ws[f'A{row}']
        cell.value = f"{t['consolidado']} ({t['empresas']}: {', '.join(map(str, params['lista_filiais']))})"
        cell.font = header_font
        cell.fill = white_fill
        row += 1
    
    cell = ws[f'G{row}']
    cell.value = f"{t['hora']}: {datetime.now().strftime('%H:%M:%S')}"
    cell.font = normal_font
    cell.alignment = right_align
    cell.fill = white_fill
    row += 2
    
    # TÍTULO - Com fundo branco
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = t['balancete']
    cell.font = Font(name='Arial', size=14, bold=True)
    cell.alignment = center_align
    cell.fill = white_fill
    row += 2
    
    # CABEÇALHOS DA TABELA - Completamente limpos (sem cores e sem bordas)
    headers = [t['codigo'], t['classificacao'], t['descricao'], t['saldo_anterior'], t['debito'], t['credito'], t['saldo_atual']]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = white_fill  # Garantir fundo branco
    
    row += 1
    
    # Congela os painéis para manter o cabeçalho visível
    ws.freeze_panes = f'A{row}'
    
    # DADOS DO BALANCETE
    linha_num = 1
    for index, data_row in df_report.iterrows():
        nivel_atual = calcular_nivel_hierarquico(str(data_row['CLAS_CTA']))
        
        # Fonte (negrito para sintéticas)
        font = bold_font if data_row['TIPO_CTA'] == 'S' else normal_font
        
        # Dados - Completamente limpos (sem bordas e com fundo branco garantido)
        cell = ws.cell(row=row, column=1, value=int(data_row['CODI_CTA']))
        cell.font = font
        cell.alignment = right_align
        cell.fill = white_fill
        
        cell = ws.cell(row=row, column=2, value=format_classificacao(data_row['CLAS_CTA']))
        cell.font = font
        cell.alignment = left_align
        cell.fill = white_fill
        
        # Nome com indentação hierárquica (SEM truncamento no XLSX)
        nome_hierarquico = format_nome_hierarquico(str(data_row['NOME_CTA']), nivel_atual)
        cell = ws.cell(row=row, column=3, value=nome_hierarquico)
        cell.font = font
        cell.alignment = left_align
        cell.fill = white_fill
        
        # Valores monetários
        cell = ws.cell(row=row, column=4, value=format_currency(data_row['SALDOANT']))
        cell.font = font
        cell.alignment = right_align
        cell.fill = white_fill
        
        cell = ws.cell(row=row, column=5, value=format_currency(data_row['TOTDEB'], show_suffix=False))
        cell.font = font
        cell.alignment = right_align
        cell.fill = white_fill
        
        cell = ws.cell(row=row, column=6, value=format_currency(data_row['TOTCRE'], show_suffix=False))
        cell.font = font
        cell.alignment = right_align
        cell.fill = white_fill
        
        cell = ws.cell(row=row, column=7, value=format_currency(data_row['SALDOATU']))
        cell.font = font
        cell.alignment = right_align
        cell.fill = white_fill
        
        row += 1
        linha_num += 1
    
    row += 1
    
    # SEÇÃO DE RESUMO - Com fundo branco
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = t['resumo_balancete']
    cell.font = Font(name='Arial', size=12, bold=True)
    cell.fill = white_fill
    row += 2
    
    # Cabeçalhos do resumo - Completamente limpos (sem cores e sem bordas)
    resume_headers = [t['categoria'], '', '', t['saldo_anterior'], t['debito'], t['credito'], t['saldo_atual']]
    for col, header in enumerate(resume_headers, 1):
        if header:  # Só preenche se não for vazio
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = bold_font
            cell.alignment = center_align
            cell.fill = white_fill  # Garantir fundo branco
    row += 1
    
    # Mapeamento de categorias para traduções
    category_translations = {
        "ATIVO": t['ativo'],
        "PASSIVO": t['passivo'], 
        "RECEITAS": t['receitas'],
        "CUSTOS DAS VENDAS": t['custos_vendas'],
        "DESPESAS OPERACIONAIS": t['despesas_operacionais'],
        "RECEITAS LIQUIDAS DOS PRODUTOS E SERVICOS - SCP": t['receitas_liquidas_scp'],
        "DESPESAS OPERACIONAIS - SCP": t['despesas_operacionais_scp'],
        "APURACAO DE RESULTADO - TRANSITORIA": t['apuracao_resultado']
    }
    
    # Calcular totais do resumo (usando apenas analíticas)
    df_analiticas = df_report[df_report['TIPO_CTA'] == 'A'].copy()
    
    summary_data = {}
    for nome, config in SUMMARY_CATEGORIES.items():
        if config['prefixes'] in [["61"], ["62"], ["9"]]:
            summary_data[nome] = {'saldo_ant': 0, 'tot_deb': 0, 'tot_cre': 0, 'saldo_atu': 0}
        else:
            mask = df_analiticas['CLAS_CTA'].str.startswith(tuple(config['prefixes']))
            df_grupo = df_analiticas[mask]
            summary_data[nome] = {
                'saldo_ant': df_grupo['SALDOANT'].sum(),
                'tot_deb': df_grupo['TOTDEB'].sum(),
                'tot_cre': df_grupo['TOTCRE'].sum(),
                'saldo_atu': df_grupo['SALDOATU'].sum()
            }
        
        # Escrever linha do resumo - Completamente limpo (sem bordas e com fundo branco)
        categoria_nome = category_translations.get(nome, nome)
        cell = ws.cell(row=row, column=1, value=categoria_nome)
        cell.font = normal_font
        cell.alignment = left_align
        cell.fill = white_fill
        
        cell = ws.cell(row=row, column=4, value=format_currency(summary_data[nome]['saldo_ant']))
        cell.font = normal_font
        cell.alignment = right_align
        cell.fill = white_fill
        
        cell = ws.cell(row=row, column=5, value=format_currency(summary_data[nome]['tot_deb'], show_suffix=False))
        cell.font = normal_font
        cell.alignment = right_align
        cell.fill = white_fill
        
        cell = ws.cell(row=row, column=6, value=format_currency(summary_data[nome]['tot_cre'], show_suffix=False))
        cell.font = normal_font
        cell.alignment = right_align
        cell.fill = white_fill
        
        cell = ws.cell(row=row, column=7, value=format_currency(summary_data[nome]['saldo_atu']))
        cell.font = normal_font
        cell.alignment = right_align
        cell.fill = white_fill
        
        row += 1
        
        # Duplicar a linha "APURACAO DE RESULTADO - TRANSITORIA"
        if nome == "APURACAO DE RESULTADO - TRANSITORIA":
            cell = ws.cell(row=row, column=1, value=categoria_nome)
            cell.font = normal_font
            cell.alignment = left_align
            cell.fill = white_fill
            
            cell = ws.cell(row=row, column=4, value=format_currency(summary_data[nome]['saldo_ant']))
            cell.font = normal_font
            cell.alignment = right_align
            cell.fill = white_fill
            
            cell = ws.cell(row=row, column=5, value=format_currency(summary_data[nome]['tot_deb'], show_suffix=False))
            cell.font = normal_font
            cell.alignment = right_align
            cell.fill = white_fill
            
            cell = ws.cell(row=row, column=6, value=format_currency(summary_data[nome]['tot_cre'], show_suffix=False))
            cell.font = normal_font
            cell.alignment = right_align
            cell.fill = white_fill
            
            cell = ws.cell(row=row, column=7, value=format_currency(summary_data[nome]['saldo_atu']))
            cell.font = normal_font
            cell.alignment = right_align
            cell.fill = white_fill
            
            row += 1
    
    row += 1
    
    # Totalizações finais do resumo (Devedoras, Credoras, Resultado)
    devedoras = {
        'saldo_ant': (summary_data['ATIVO']['saldo_ant'] + summary_data['CUSTOS DAS VENDAS']['saldo_ant'] + summary_data['DESPESAS OPERACIONAIS']['saldo_ant']),
        'tot_deb': (summary_data['ATIVO']['tot_deb'] + summary_data['CUSTOS DAS VENDAS']['tot_deb'] + summary_data['DESPESAS OPERACIONAIS']['tot_deb']),
        'tot_cre': (summary_data['ATIVO']['tot_cre'] + summary_data['CUSTOS DAS VENDAS']['tot_cre'] + summary_data['DESPESAS OPERACIONAIS']['tot_cre']),
        'saldo_atu': (summary_data['ATIVO']['saldo_atu'] + summary_data['CUSTOS DAS VENDAS']['saldo_atu'] + summary_data['DESPESAS OPERACIONAIS']['saldo_atu'])
    }
    
    credoras = {
        'saldo_ant': summary_data['PASSIVO']['saldo_ant'] + summary_data['RECEITAS']['saldo_ant'],
        'tot_deb': summary_data['PASSIVO']['tot_deb'] + summary_data['RECEITAS']['tot_deb'],
        'tot_cre': summary_data['PASSIVO']['tot_cre'] + summary_data['RECEITAS']['tot_cre'],
        'saldo_atu': summary_data['PASSIVO']['saldo_atu'] + summary_data['RECEITAS']['saldo_atu']
    }
    
    # CONTAS DEVEDORAS - Completamente limpo
    cell = ws.cell(row=row, column=1, value=t['contas_devedoras'])
    cell.font = bold_font
    cell.fill = white_fill
    cell = ws.cell(row=row, column=4, value=format_currency(devedoras['saldo_ant']))
    cell.font = bold_font
    cell.alignment = right_align
    cell.fill = white_fill
    cell = ws.cell(row=row, column=5, value=format_currency(devedoras['tot_deb'], show_suffix=False))
    cell.font = bold_font
    cell.alignment = right_align
    cell.fill = white_fill
    cell = ws.cell(row=row, column=6, value=format_currency(devedoras['tot_cre'], show_suffix=False))
    cell.font = bold_font
    cell.alignment = right_align
    cell.fill = white_fill
    cell = ws.cell(row=row, column=7, value=format_currency(devedoras['saldo_atu']))
    cell.font = bold_font
    cell.alignment = right_align
    cell.fill = white_fill
    row += 1
    
    # CONTAS CREDORAS - Completamente limpo
    cell = ws.cell(row=row, column=1, value=t['contas_credoras'])
    cell.font = bold_font
    cell.fill = white_fill
    cell = ws.cell(row=row, column=4, value=format_currency(credoras['saldo_ant']))
    cell.font = bold_font
    cell.alignment = right_align
    cell.fill = white_fill
    cell = ws.cell(row=row, column=5, value=format_currency(credoras['tot_deb'], show_suffix=False))
    cell.font = bold_font
    cell.alignment = right_align
    cell.fill = white_fill
    cell = ws.cell(row=row, column=6, value=format_currency(credoras['tot_cre'], show_suffix=False))
    cell.font = bold_font
    cell.alignment = right_align
    cell.fill = white_fill
    cell = ws.cell(row=row, column=7, value=format_currency(credoras['saldo_atu']))
    cell.font = bold_font
    cell.alignment = right_align
    cell.fill = white_fill
    row += 1
    
    # Cálculos de resultado
    receitas_analiticas = df_analiticas[df_analiticas['CLAS_CTA'].str.startswith('3')]
    custos_analiticas = df_analiticas[df_analiticas['CLAS_CTA'].str.startswith('4')]
    despesas_analiticas = df_analiticas[df_analiticas['CLAS_CTA'].str.startswith('5')]
    
    debito_total_mes = (custos_analiticas['TOTDEB'].sum() - custos_analiticas['TOTCRE'].sum()) + (despesas_analiticas['TOTDEB'].sum() - despesas_analiticas['TOTCRE'].sum())
    credito_total_mes = receitas_analiticas['TOTCRE'].sum() - receitas_analiticas['TOTDEB'].sum()
    resultado_mes = credito_total_mes - debito_total_mes
    
    saldo_anterior_exercicio = (custos_analiticas['SALDOANT'].sum() + despesas_analiticas['SALDOANT'].sum()) + receitas_analiticas['SALDOANT'].sum()
    debito_exercicio = custos_analiticas['SALDOATU'].sum() + despesas_analiticas['SALDOATU'].sum()
    credito_exercicio = abs(receitas_analiticas['SALDOATU'].sum())
    resultado_exercicio = credito_exercicio - debito_exercicio
    
    # RESULTADO DO MÊS - Completamente limpo
    cell = ws.cell(row=row, column=1, value=t['resultado_mes'])
    cell.font = bold_font
    cell.fill = white_fill
    cell = ws.cell(row=row, column=4, value="-")
    cell.font = bold_font
    cell.alignment = center_align
    cell.fill = white_fill
    cell = ws.cell(row=row, column=5, value=format_currency(debito_total_mes, show_suffix=False))
    cell.font = bold_font
    cell.alignment = right_align
    cell.fill = white_fill
    cell = ws.cell(row=row, column=6, value=format_currency(credito_total_mes, show_suffix=False))
    cell.font = bold_font
    cell.alignment = right_align
    cell.fill = white_fill
    cell = ws.cell(row=row, column=7, value=format_currency_resultado(resultado_mes))
    cell.font = bold_font
    cell.alignment = right_align
    cell.fill = white_fill
    row += 1
    
    # RESULTADO DO EXERCÍCIO - Completamente limpo
    cell = ws.cell(row=row, column=1, value=t['resultado_exercicio'])
    cell.font = bold_font
    cell.fill = white_fill
    cell = ws.cell(row=row, column=4, value=format_currency(saldo_anterior_exercicio))
    cell.font = bold_font
    cell.alignment = right_align
    cell.fill = white_fill
    cell = ws.cell(row=row, column=5, value=format_currency(debito_exercicio, show_suffix=False))
    cell.font = bold_font
    cell.alignment = right_align
    cell.fill = white_fill
    cell = ws.cell(row=row, column=6, value=format_currency(credito_exercicio, show_suffix=False))
    cell.font = bold_font
    cell.alignment = right_align
    cell.fill = white_fill
    cell = ws.cell(row=row, column=7, value=format_currency_resultado(resultado_exercicio))
    cell.font = bold_font
    cell.alignment = right_align
    cell.fill = white_fill
    row += 3
    
    # RODAPÉ COM ASSINATURAS - Completamente limpo
    cell = ws.cell(row=row, column=2, value="_" * 40)
    cell.font = normal_font
    cell.alignment = center_align
    cell.fill = white_fill
    cell = ws.cell(row=row, column=6, value="_" * 40)
    cell.font = normal_font
    cell.alignment = center_align
    cell.fill = white_fill
    row += 1
    
    cell = ws.cell(row=row, column=2, value=footer_data['rep_legal_nome'])
    cell.font = normal_font
    cell.alignment = center_align
    cell.fill = white_fill
    cell = ws.cell(row=row, column=6, value=footer_data['contador_nome'])
    cell.font = normal_font
    cell.alignment = center_align
    cell.fill = white_fill
    row += 1
    
    cell = ws.cell(row=row, column=2, value=f"CPF: {footer_data['rep_legal_cpf']}")
    cell.font = normal_font
    cell.alignment = center_align
    cell.fill = white_fill
    cell = ws.cell(row=row, column=6, value=f"Reg. no CRC - {footer_data['contador_uf_crc']} sob o No. {footer_data['contador_crc']}")
    cell.font = normal_font
    cell.alignment = center_align
    cell.fill = white_fill
    row += 1
    
    cell = ws.cell(row=row, column=6, value=f"CPF: {footer_data['contador_cpf']}")
    cell.font = normal_font
    cell.alignment = center_align
    cell.fill = white_fill
    
    # Ajustar largura das colunas
    column_widths = [10, 15, 40, 15, 15, 15, 15]  # Larguras aproximadas
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Salvar arquivo
    wb.save(filename)
    print(f"   XLSX '{filename}' gerado com sucesso!")
    return filename

def GerarRelatorioBalancete(codi_emp, data_inicial, data_final, ingles=False):
    """
    Função principal para gerar relatórios de balancete.
    
    Args:
        codi_emp (int): Código da empresa principal
        data_inicial (str): Data inicial no formato 'YYYY-MM-DD'
        data_final (str): Data final no formato 'YYYY-MM-DD'  
        ingles (bool): Se True, gera relatório em inglês usando traduções do banco
    
    Returns:
        dict: Um dicionário com os nomes dos arquivos gerados {'pdf': ..., 'xlsx': ...}
              ou None em caso de falha.
    """
    conn = None
    params = None
    try:
        print("=" * 60)
        print("🏢 SISTEMA DE GERAÇÃO DE BALANCETE CONTÁBIL")
        print("=" * 60)
        
        # Gera um timestamp único para esta execução
        timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Validar datas
        datetime.strptime(data_inicial, '%Y-%m-%d')
        datetime.strptime(data_final, '%Y-%m-%d')

        print(f"\n📋 PARÂMETROS CONFIGURADOS:")
        print(f"   • Empresa: {codi_emp}")
        print(f"   • Período: {data_inicial} a {data_final}")
        print(f"   • Idioma: {'Inglês' if ingles else 'Português'}")
        print(f"   • Assinatura de Arquivo: {timestamp_str}")

        # Buscar parâmetros
        print("\n🔍 Buscando parâmetros de configuração...")
        params = obter_parametros_relatorio_balancete(codi_emp)
        if not params:
            print("❌ Não foi possível obter os parâmetros de configuração. Abortando.")
            return None
        params['empresa_principal'] = codi_emp

        print("\n⚡ Gerando consulta SQL otimizada...")
        consulta_otimizada = gerar_consulta_otimizada(params, data_inicial, data_final, ingles)
        
        print("\n🔄 Executando consulta única consolidada...")
        start_time = datetime.now()
        conn = pyodbc.connect(CONN_STR)
        
        print("📊 Buscando dados do relatório...")
        df_report = pd.read_sql(consulta_otimizada, conn)
        
        query_time = datetime.now()
        print(f"   ✅ {len(df_report)} linhas brutas encontradas em {(query_time - start_time).total_seconds():.2f}s")
        
        if df_report.empty:
            print("\n⚠️ AVISO: Nenhuma linha retornada na consulta. O relatório não será gerado.")
            return None
        
        sinteticas_count = len(df_report[df_report['TIPO_CTA'] == 'S'])
        analiticas_count = len(df_report[df_report['TIPO_CTA'] == 'A'])
        print(f"   📊 Composição inicial: {sinteticas_count} sintéticas + {analiticas_count} analíticas")

        df_report = aplicar_filtros_negocios(df_report)
        
        if df_report.empty:
            print("\n⚠️ AVISO: Nenhuma linha restou após os filtros. O relatório não será gerado.")
            return None

        print("📋 Buscando dados de cabeçalho e rodapé...")
        header_data, footer_data = buscar_dados_auxiliares(codi_emp)
        
        filtros_time = datetime.now()
        
        print("🧮 Calculando totalização hierárquica...")
        hierarchy_start = datetime.now()
        df_report = calcular_totalizacao_hierarquica_otimizada(df_report)
        hierarchy_time = datetime.now()
        print(f"   ✅ Hierarquia calculada em {(hierarchy_time - hierarchy_start).total_seconds():.2f}s")
        
        print("🗑️ Filtrando contas sintéticas zeradas sem descendentes...")
        contas_sinteticas = df_report[df_report['TIPO_CTA'] == 'S'].copy()
        contas_analiticas = df_report[df_report['TIPO_CTA'] == 'A'].copy()
        clas_analiticas_set = set(contas_analiticas['CLAS_CTA'].astype(str))
        
        indices_para_remover = []
        for idx, conta in contas_sinteticas.iterrows():
            if (conta['SALDOANT'] == 0 and conta['TOTDEB'] == 0 and 
                conta['TOTCRE'] == 0 and conta['SALDOATU'] == 0):
                
                clas_sintetica = str(conta['CLAS_CTA']).strip()
                tem_descendentes = False
                for clas_analitica in clas_analiticas_set:
                    if any(clas_analitica.startswith(clas_sintetica + str(i)) for i in range(10)):
                        analitic_data = contas_analiticas[contas_analiticas['CLAS_CTA'] == clas_analitica]
                        if not analitic_data.empty:
                            row_data = analitic_data.iloc[0]
                            if (row_data['SALDOANT'] != 0 or row_data['TOTDEB'] != 0 or 
                                row_data['TOTCRE'] != 0 or row_data['SALDOATU'] != 0):
                                tem_descendentes = True
                                break
                
                if not tem_descendentes:
                    indices_para_remover.append(idx)
        
        if indices_para_remover:
            df_report = df_report.drop(indices_para_remover)
            print(f"   ✅ {len(indices_para_remover)} contas sintéticas zeradas removidas.")
        
        conn.commit()
        print("   ✅ Transação confirmada (commit).")
        
        print("\n📄 Gerando relatórios de saída...")
        generation_start = datetime.now()
        
        pdf_filename = ""
        xlsx_filename = ""
        
        with ThreadPoolExecutor(max_workers=2) as executor:
            pdf_future = executor.submit(generate_pdf, df_report, header_data, footer_data, params, data_inicial, data_final, timestamp_str, ingles)
            xlsx_future = executor.submit(generate_xlsx, df_report, header_data, footer_data, params, data_inicial, data_final, timestamp_str, ingles)
            
            pdf_filename = pdf_future.result()
            xlsx_filename = xlsx_future.result()
        
        generation_time = datetime.now()
        total_time = generation_time - start_time
        
        print(f"\n🎉 === PROCESSO CONCLUÍDO COM SUCESSO ===")
        print(f"⏱️  Tempo total: {total_time.total_seconds():.2f} segundos")
        print(f"📊 Consulta SQL: {(query_time - start_time).total_seconds():.2f}s")
        print(f"🔍 Filtros de negócio: {(filtros_time - query_time).total_seconds():.2f}s")
        print(f"🧮 Hierarquia: {(hierarchy_time - hierarchy_start).total_seconds():.2f}s")
        print(f"📄 Geração de arquivos: {(generation_time - generation_start).total_seconds():.2f}s")
        print(f"📁 Arquivos gerados:")
        print(f"   • {pdf_filename}")
        print(f"   • {xlsx_filename}")

        # Renomear arquivos conforme padrão solicitado
        print(f"\n📝 Renomeando arquivos...")
        try:
            pdf_path_renamed = rename_file_after_generation(
                pdf_filename, str(params['empresa_principal']), data_inicial, data_final,
                timestamp_str, ingles, 'balancete', CONN_STR
            )
            xlsx_path_renamed = rename_file_after_generation(
                xlsx_filename, str(params['empresa_principal']), data_inicial, data_final,
                timestamp_str, ingles, 'balancete', CONN_STR
            )
            pdf_filename = pdf_path_renamed
            xlsx_filename = xlsx_path_renamed
        except Exception as e:
            print(f"⚠️  Aviso: Erro ao renomear arquivos: {e}")

        return {
            "pdf": pdf_filename,
            "xlsx": xlsx_filename
        }

    except pyodbc.Error as ex:
        print(f"\n❌ ERRO DE BANCO DE DADOS: {ex.args[0]}\n{ex}")
        return None
    except ValueError as e:
        print(f"\n❌ ERRO: Entrada de dados inválida. {e}")
        return None
    except Exception as e:
        print(f"\n❌ Ocorreu um erro inesperado: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        if conn:
            print("🔌 Fechando a conexão com o banco de dados.")
            conn.close()

